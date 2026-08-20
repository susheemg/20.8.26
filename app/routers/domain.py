"""Auto-extracted domain routes (RouterDeps pattern). See app/routers/deps.py.

Behaviour is byte-identical to the pre-split monolith; per-instance deps are bound
as locals (multi-app isolation), invariant models/imports come from bro_app globals.
"""
from __future__ import annotations

from fastapi import APIRouter, Request
import app.features.admin.rbac as _RBAC
from sqlalchemy import or_ as _or2
import json as _json2
from fastapi.responses import (PlainTextResponse, StreamingResponse,
    HTMLResponse, JSONResponse, FileResponse, RedirectResponse)

from .deps import RouterDeps
from ._shared import bind_shared


def build_domain_router(deps: RouterDeps) -> APIRouter:
    import app.bro_app as _M
    globals().update({k: v for k, v in vars(_M).items() if not k.startswith("__")})
    r = APIRouter()
    app = r
    db = deps.db
    actor = deps.actor
    require = deps.require
    audit = deps.audit
    notify = deps.notify
    _fb_guidance = deps.fb_guidance
    _ai_live = deps.ai_live
    AI_HOLDING = deps.ai_holding
    engine = deps.engine
    _platform_version = deps.platform_version
    SessionFactory = deps.session_factory
    _sh = bind_shared(deps)
    _monitor_interval = _sh["_monitor_interval"]
    _rmd_row = _sh["_rmd_row"]
    _file_monitoring_report = _sh["_file_monitoring_report"]
    _ai_research = _sh["_ai_research"]
    _start_research_job = _sh["_start_research_job"]
    _research_job_status = _sh["_research_job_status"]
    # --- build-level imports replicated from the pre-split factory ---
    from app.features.admin import identity as IDP
    from app.features.domain.models_db import hash_password as _hash_pw
    import secrets as _secrets
    from app.features.intelligence import sanctions as SANC
    from app.features.domain.registry_models import SanctionsScreening, WatchlistEntry, VendorPerson
    from app.features.lifecycle import monitoring as MON
    from fastapi.responses import PlainTextResponse
    import csv as _csv, io as _io
    from datetime import datetime as _dt2
    from app.features.assessment import agents as _A
    from app.features.assessment import agent_engine as _AE
    from app.features.domain.models_feature import AgentLearning, BackgroundInsight
    from app.features.domain import registry_service as RS
    from app.features.intelligence import financial as FIN
    from app.features.domain.registry_models import (
        IndustryMaster, MaterialGroupMaster, VendorGroup, VendorRecord,
        VendorIndustry, ContactRecord, EngagementRecord, AssessmentRecord,
        FindingRecord, RemediationRecord, FourthPartyRecord, FourthPartyVendor,
        ArtefactRecord, IssueRecord,
    )
    import json as _json2, os as _os2
    from app.features.lifecycle import performance_service as PERF
    from app.features.platform import platform_docs as PDOCS
    from app.features.assessment import learnings as LEARN
    from app.features.admin import integrations as INTEG
    from app.features.admin import content as CONTENT
    from app.features.admin import layout as LAYOUT


    @app.post("/api/v1/vendors")
    def create_vendor(v: VendorIn, s: Session = Depends(db),
                      u: User = Depends(require("vendor.edit"))):
        row = Vendor(**v.model_dump())
        s.add(row); s.flush()
        audit(s, "vendor.created", u.username, {"vendor_id": row.id, "name": v.name})
        s.commit()
        return {"vendor_id": row.id, "name": row.name, "tier": row.tier}

    @app.get("/api/v1/vendors")
    def list_vendors(limit: Optional[int] = None, offset: int = 0,
                     s: Session = Depends(db), u: User = Depends(require("vendor.view"))):
        q = select(Vendor).offset(max(0, offset))
        if limit:
            q = q.limit(max(1, min(int(limit), 500)))
        return [{"vendor_id": v.id, "name": v.name, "tier": v.tier,
                 "is_critical": v.is_critical}
                for v in s.scalars(q).all()]

    @app.post("/api/v1/vendors/{vid}/critical")
    def designate_critical(vid: int, body: CriticalIn, s: Session = Depends(db),
                           u: User = Depends(require("vendor.critical"))):
        _RBAC.assert_object_visible(s, u, 'vendor', vid)
        # Tier 0 = human-only (our Q5). RBAC already enforces a human role here.
        v = s.get(Vendor, vid)
        if not v:
            raise HTTPException(404, "vendor not found")
        v.is_critical = True
        v.critical_reason = body.reason
        v.critical_by = u.username
        audit(s, "vendor.critical_designated", u.username,
              {"vendor_id": vid, "reason": body.reason})
        notify(s, f"Critical vendor designated: {v.name}", "vrm")
        s.commit()
        return {"vendor_id": vid, "is_critical": True, "by": u.username}

    # ===== engagements + lifecycle =====
    @app.post("/api/v1/engagements")
    def create_engagement(e: EngagementIn, s: Session = Depends(db),
                          u: User = Depends(require("engagement.create"))):
        if not s.get(Vendor, e.vendor_id):
            raise HTTPException(404, "vendor not found")
        row = EngagementRow(vendor_id=e.vendor_id, title=e.title,
                            service_description=e.service_description,
                            business_contact_email=e.business_contact_email,
                            owner_id=u.id, stage="sourcing")
        s.add(row); s.flush()
        audit(s, "engagement.created", u.username, {"engagement_id": row.id})
        notify(s, f"Engagement created: {e.title}", "business")
        s.commit()
        return {"engagement_id": row.id, "stage": row.stage}

    @app.post("/api/v1/engagements/{eid}/irq")
    def submit_irq(eid: int, body: IRQIn, s: Session = Depends(db),
                   u: User = Depends(require("engagement.edit"))):
        _RBAC.assert_object_visible(s, u, 'engagement', eid)
        e = s.get(EngagementRow, eid)
        if not e:
            raise HTTPException(404, "engagement not found")
        tier = eng.compute_tier(body.answers)
        inherent = eng.compute_inherent(body.answers)
        routing = eng.compute_route(body.answers, inherent, tier)
        e.inherent_band = inherent["band"]
        e.inherent_pct = inherent["weighted_pct"]
        e.route = routing["route"]
        e.stage = "inherent"
        audit(s, "irq.scored", u.username,
              {"engagement_id": eid, "tier": tier, "band": inherent["band"],
               "route": routing["route"]})
        notify(s, f"IRQ scored {inherent['band']} ({routing['route']})", "all")
        s.commit()
        return {"engagement_id": eid, "tier": tier,
                "inherent_band": inherent["band"],
                "inherent_pct": inherent["weighted_pct"],
                "cls": inherent["cls"], "routing": routing}

    @app.post("/api/v1/engagements/{eid}/ddq")
    def submit_ddq(eid: int, body: DDQIn, s: Session = Depends(db),
                   u: User = Depends(require("engagement.edit"))):
        _RBAC.assert_object_visible(s, u, 'engagement', eid)
        e = s.get(EngagementRow, eid)
        if not e:
            raise HTTPException(404, "engagement not found")
        residual = eng.compute_residual(e.inherent_band or "LOW", body.answers)
        decision = eng.decision_for(residual["band"], residual["critical_marginal"])
        e.residual_band = residual["band"]
        e.decision = decision["text"]
        e.stage = "decision"
        audit(s, "ddq.residual", u.username,
              {"engagement_id": eid, "residual": residual["band"],
               "critical_marginal": residual["critical_marginal"],
               "decision": decision["text"]})
        notify(s, f"Residual {residual['band']}: {decision['text']}", "all")
        s.commit()
        return {"engagement_id": eid, "residual_band": residual["band"],
                "critical_marginal": residual["critical_marginal"],
                "decision": decision}

    @app.post("/api/v1/engagements/{eid}/override")
    def override(eid: int, body: OverrideIn, s: Session = Depends(db),
                 u: User = Depends(require("engagement.override"))):
        _RBAC.assert_object_visible(s, u, 'engagement', eid)
        # human-only (RBAC) + justification + 2nd approver (two-gate model)
        e = s.get(EngagementRow, eid)
        if not e:
            raise HTTPException(404, "engagement not found")
        if not body.reason or not body.second_approver:
            raise HTTPException(400, "override needs justification and 2nd approver")
        e.residual_band = body.band
        audit(s, "decision.override", u.username,
              {"engagement_id": eid, "new_band": body.band,
               "reason": body.reason, "second_approver": body.second_approver})
        s.commit()
        return {"engagement_id": eid, "residual_band": body.band, "override": True}

    @app.post("/api/v1/engagements/{eid}/terminate")
    def terminate(eid: int, s: Session = Depends(db),
                  u: User = Depends(require("lifecycle.offboard"))):
        _RBAC.assert_object_visible(s, u, 'engagement', eid)
        e = s.get(EngagementRow, eid)
        if not e:
            raise HTTPException(404, "engagement not found")
        e.stage = "terminate"
        e.status = "terminated"
        for key, label in eng.OFFBOARDING_STEPS:
            s.add(Offboarding(engagement_id=eid, step_key=key))
        audit(s, "engagement.terminated", u.username, {"engagement_id": eid})
        notify(s, "Offboarding initiated", "all")
        s.commit()
        return {"engagement_id": eid, "stage": "terminate",
                "offboarding_steps": len(eng.OFFBOARDING_STEPS)}

    @app.get("/api/v1/engagements/{eid}")
    def get_engagement(eid: int, s: Session = Depends(db),
                       u: User = Depends(require("engagement.view"))):
        _RBAC.assert_object_visible(s, u, 'engagement', eid)
        e = s.get(EngagementRow, eid)
        if not e:
            raise HTTPException(404, "engagement not found")
        return {"engagement_id": e.id, "vendor_id": e.vendor_id,
                "stage": e.stage, "status": e.status, "route": e.route,
                "inherent_band": e.inherent_band, "residual_band": e.residual_band,
                "decision": e.decision}

    # ===== findings =====
    @app.get("/api/v2/vendors/{vid}/people")
    def v2_vendor_people(vid: str, s: Session = Depends(db),
                         u: User = Depends(require("vendor.view"))):
        _RBAC.assert_object_visible(s, u, 'vendor', vid)
        rows = s.scalars(select(VendorPerson).where(VendorPerson.vendor_id == vid)).all()
        return [{"person_id": p.person_id, "name": p.name, "role": p.role, "dob": p.dob,
                 "nationality": p.nationality, "ownership_pct": p.ownership_pct,
                 "is_ubo": p.is_ubo} for p in rows]

    @app.post("/api/v2/vendors/{vid}/people")
    def v2_add_vendor_person(vid: str, body: dict = Body(...), s: Session = Depends(db),
                             u: User = Depends(require("vendor.edit"))):
        _RBAC.assert_object_visible(s, u, 'vendor', vid)
        pid = RS.next_id(s, "person")
        s.add(VendorPerson(person_id=pid, vendor_id=vid, name=body.get("name", "").strip(),
                           role=body.get("role"), dob=body.get("dob"),
                           nationality=body.get("nationality"),
                           ownership_pct=body.get("ownership_pct"),
                           is_ubo=bool(body.get("is_ubo"))))
        audit(s, "vendor.person_added", u.username, {"vendor_id": vid, "person_id": pid})
        s.commit()
        return {"person_id": pid}

    @app.post("/api/v1/assess/start")
    def assess_start(b: ChatStart, s: Session = Depends(db),
                     u: User = Depends(require("engagement.view"))):
        sess = ConversationSession(engagement_id=b.engagement_id,
                                   actor_role=b.actor_role)
        s.add(sess); s.flush()
        s.commit()
        return {"session_id": sess.id, "actor_role": sess.actor_role}

    @app.post("/api/v1/assess/turn")
    def assess_turn(b: ChatTurn, s: Session = Depends(db),
                    u: User = Depends(require("engagement.view"))):
        sess = s.get(ConversationSession, b.session_id)
        if not sess:
            raise HTTPException(404, "session not found")
        s.add(ConversationMessage(session_id=b.session_id,
                                  role=sess.actor_role, body=b.message))
        # deterministic adaptive reply: vendor claims are flagged to verify;
        # assessor input is trusted (our role/trust model)
        if sess.actor_role == "vendor":
            reply = ("Noted as a vendor assertion — this will be verified against "
                     "independent evidence before it affects the rating.")
            visibility = "shared"
        else:
            reply = "Recorded as assessor input and applied to the assessment."
            visibility = "internal"
        s.add(ConversationMessage(session_id=b.session_id, role="assistant", body=reply))
        s.commit()
        return {"session_id": b.session_id, "reply": reply, "visibility": visibility}

    # ===== autopilot (propose, human executes — two-gate) =====
    @app.post("/api/v1/engagements/{eid}/autopilot")
    def autopilot(eid: int, body: IRQIn, s: Session = Depends(db),
                  u: User = Depends(require("engagement.autopilot"))):
        _RBAC.assert_object_visible(s, u, 'engagement', eid)
        e = s.get(EngagementRow, eid)
        if not e:
            raise HTTPException(404, "engagement not found")
        tier = eng.compute_tier(body.answers)
        inherent = eng.compute_inherent(body.answers)
        routing = eng.compute_route(body.answers, inherent, tier)
        # Watchlist gate: a watchlisted supplier forces human sign-off regardless of band.
        from app.features.domain import watchlist_service as _WL
        wl_signoff = _WL.is_watchlisted(s, e.vendor_id) if getattr(e, "vendor_id", None) else False
        # PROPOSES — does not finalise the decision; a human must record it.
        audit(s, "autopilot.proposed", u.username,
              {"engagement_id": eid, "proposed_band": inherent["band"],
               "watchlist_signoff": wl_signoff})
        s.commit()
        return {"engagement_id": eid, "proposed_tier": tier,
                "proposed_inherent": inherent, "proposed_routing": routing,
                "watchlist_signoff_required": wl_signoff,
                "status": ("PROPOSED — supplier on watchlist: human sign-off is mandatory "
                           "irrespective of risk level." if wl_signoff
                           else "PROPOSED — requires human to record decision")}

    # ===== notifications =====
    @app.get("/api/v1/audit")
    def audit_trail(s: Session = Depends(db), u: User = Depends(require("audit.view")),
                    limit: int = 200, before_seq: int = 0,
                    vendor_id: str = "", engagement_id: str = "",
                    actor: str = "", action: str = ""):
        """Paginated, filterable audit trail (DB-03).

        Previously loaded the entire table. Now returns a bounded page, newest first,
        with optional subject/actor/action filters served by the new indexes."""
        lim = min(max(int(limit or 200), 1), 500)
        q = select(AuditLog)
        if before_seq:
            q = q.where(AuditLog.seq < int(before_seq))
        if vendor_id:
            q = q.where(AuditLog.vendor_id == vendor_id)
        if engagement_id:
            q = q.where(AuditLog.engagement_id == engagement_id)
        if actor:
            q = q.where(AuditLog.actor == actor)
        if action:
            q = q.where(AuditLog.action == action)
        rows = s.scalars(q.order_by(AuditLog.seq.desc()).limit(lim + 1)).all()
        more = len(rows) > lim
        rows = rows[:lim]
        return {"entries": [{"seq": r.seq, "action": r.action, "actor": r.actor,
                             "vendor_id": r.vendor_id, "engagement_id": r.engagement_id,
                             "timestamp": r.created_at.isoformat() if r.created_at else None,
                             "hash": r.entry_hash} for r in rows],
                "count": len(rows), "has_more": more,
                "next_before_seq": (rows[-1].seq if rows and more else None)}

    @app.get("/api/v1/audit/verify")
    def audit_verify(s: Session = Depends(db), u: User = Depends(require("audit.view")),
                     full: int = 0, window: int = 5000):
        """Verify the hash chain (DB-03).

        Walking the whole chain on every call does not scale, so by default this
        verifies the most recent `window` entries and confirms they link back to a
        verified predecessor. `full=1` still walks the entire chain for periodic
        assurance runs and for export evidence."""
        total = s.scalar(select(func.count()).select_from(AuditLog)) or 0
        q = select(AuditLog).order_by(AuditLog.seq)
        checked_from = 0
        if not full and total > window:
            checked_from = total - window
            q = q.where(AuditLog.seq >= checked_from)
        rows = s.scalars(q).all()
        prev = rows[0].prev_hash if (rows and checked_from) else "genesis"
        for r in rows:
            expect = eng.chain_hash(prev, r.action, r.actor,
                                    json.loads(r.detail) if r.detail else {})
            if r.prev_hash != prev or r.entry_hash != expect:
                return {"intact": False, "broke_at": r.seq, "entries_checked": len(rows),
                        "scope": "full" if full else "window"}
            prev = r.entry_hash
        return {"intact": True, "entries_checked": len(rows), "total_entries": total,
                "scope": "full" if full else f"latest {len(rows)}",
                "note": None if (full or total <= window) else
                        "Windowed verification. Call with full=1 for a whole-chain walk."}

    # ===== MCP-style read tools (Group G) =====
    @app.get("/api/v1/mcp/portfolio-summary")
    def portfolio_summary(s: Session = Depends(db), u: User = Depends(require("vendor.view"))):
        vendors = s.scalars(select(Vendor)).all()
        engagements = s.scalars(select(EngagementRow)).all()
        return {"vendors": len(vendors),
                "critical_vendors": sum(1 for v in vendors if v.is_critical),
                "engagements": len(engagements),
                "by_decision": _count_by(engagements, "decision")}

    @app.get("/api/v1/mcp/critical-vendors")
    def critical_vendors(s: Session = Depends(db), u: User = Depends(require("vendor.view"))):
        return [{"vendor_id": v.id, "name": v.name, "reason": v.critical_reason}
                for v in s.scalars(select(Vendor).where(Vendor.is_critical == True)).all()]  # noqa: E712

    @app.get("/api/v1/mcp/overdue-findings")
    def overdue_findings(s: Session = Depends(db), u: User = Depends(require("finding.view"))):
        rows = s.scalars(select(Finding).where(Finding.status != "closed")).all()
        return [{"finding_id": f.id, "title": f.title, "severity": f.severity,
                 "status": f.status} for f in rows]

    # ===== procurement integration (Group G) =====

    @app.post("/api/v1/procurement/po")
    def procurement_po(b: POIn, s: Session = Depends(db),
                       u: User = Depends(require("admin.integrations"))):
        # inbound PO auto-creates a vendor + sourcing engagement (straight-through)
        v = Vendor(name=b.vendor_name, ext_ref=b.ext_ref)
        s.add(v); s.flush()
        e = EngagementRow(vendor_id=v.id, title=f"PO {b.ext_ref or v.id}",
                          stage="sourcing")
        s.add(e); s.flush()
        audit(s, "procurement.po_ingested", u.username,
              {"vendor_id": v.id, "engagement_id": e.id, "amount": b.amount})
        s.commit()
        return {"vendor_id": v.id, "engagement_id": e.id, "stage": "sourcing"}

    # ===== certifications =====
    @app.post("/api/v1/certifications")
    def add_cert(b: CertIn, s: Session = Depends(db),
                 u: User = Depends(require("lifecycle.certs"))):
        from datetime import datetime as _dt
        from app.features.domain.models_db import Vendor as _V
        vu = _dt.fromisoformat(b.valid_until) if b.valid_until else None
        # supersede any prior current cert of the same name for this vendor
        priors = s.scalars(select(Certification).where(
            Certification.vendor_id == b.vendor_id, Certification.name == b.name,
            Certification.superseded == False)).all()  # noqa: E712
        for p in priors:
            p.superseded = True
        row = Certification(vendor_id=b.vendor_id, name=b.name, valid_until=vu)
        s.add(row); s.flush()
        # Write through to the canonical artefact pipeline so the revalidation engine,
        # expiring view, Issues Log, supersession and auto-close all cover register certs.
        reg_vid = None
        v1 = s.get(_V, b.vendor_id)
        if v1:
            vr = s.scalars(select(VendorRecord).where(VendorRecord.legal_name == v1.name)).first()
            reg_vid = vr.vendor_id if vr else (v1.ext_ref or f"VEN-V1-{v1.id}")
        if reg_vid:
            prior_art = s.scalars(select(ArtefactRecord).where(
                ArtefactRecord.vendor_id == reg_vid, ArtefactRecord.name == b.name,
                ArtefactRecord.is_current == True)).first()  # noqa: E712
            art = RS.create_artefact(
                s, vendor_id=reg_vid, name=b.name, artefact_type="certificate",
                expiry_date=(vu.date().isoformat() if vu else None), is_dated=bool(vu),
                received_via="register", supersedes=prior_art.artefact_id if prior_art else None)
            row.artefact_id = art.artefact_id
        audit(s, "cert.added", u.username, {"cert_id": row.id, "vendor_id": b.vendor_id,
                                            "artefact_id": row.artefact_id})
        s.commit()
        return {"cert_id": row.id, "name": row.name, "artefact_id": row.artefact_id}

    @app.get("/api/v1/vendors/{vid}/certifications")
    def list_certs(vid: int, s: Session = Depends(db),
                   u: User = Depends(require("lifecycle.certs"))):
        _RBAC.assert_object_visible(s, u, 'vendor', vid)
        rows = s.scalars(select(Certification).where(Certification.vendor_id == vid)).all()
        return [{"cert_id": c.id, "name": c.name,
                 "valid_until": c.valid_until.isoformat() if c.valid_until else None}
                for c in rows]

    # ===== documents + evidence expiry =====
    @app.get("/api/v1/evidence/expiring")
    def expiring_evidence(s: Session = Depends(db),
                          u: User = Depends(require("lifecycle.evidence"))):
        from datetime import datetime as _dt, timedelta
        horizon = _dt.utcnow() + timedelta(days=90)
        out = []
        for d in s.scalars(select(Document).where(
                Document.next_validation != None,  # noqa: E711
                Document.next_validation <= horizon)).all():
            out.append({"source": "document", "document_id": d.id, "name": d.name,
                        "next_validation": d.next_validation.isoformat()})
        # Join the certifications register so register-added certs surface here too.
        for c in s.scalars(select(Certification).where(
                Certification.valid_until != None,  # noqa: E711
                Certification.valid_until <= horizon,
                Certification.superseded == False)).all():  # noqa: E712
            out.append({"source": "certification", "cert_id": c.id, "name": c.name,
                        "next_validation": c.valid_until.isoformat()})
        out.sort(key=lambda r: r["next_validation"])
        return out

    # ===== fourth parties + concentration =====
    @app.post("/api/v1/fourth-parties")
    def add_fourth(b: FourthIn, s: Session = Depends(db),
                   u: User = Depends(require("lifecycle.fourthparty"))):
        row = FourthParty(vendor_id=b.vendor_id, name=b.name, service=b.service)
        s.add(row); s.flush()
        # concentration: same 4th party serving many vendors
        count = s.scalar(select(func.count()).select_from(FourthParty)
                         .where(FourthParty.name == b.name))
        if count and count >= 3:
            row.concentration_flag = True
            notify(s, f"Concentration risk: {b.name} serves {count} vendors", "vrm")
        audit(s, "fourthparty.added", u.username, {"id": row.id})
        s.commit()
        return {"fourth_party_id": row.id, "concentration_flag": row.concentration_flag}

    @app.get("/api/v1/fourth-parties/concentration")
    def concentration(s: Session = Depends(db),
                      u: User = Depends(require("lifecycle.fourthparty"))):
        rows = s.scalars(select(FourthParty).where(
            FourthParty.concentration_flag == True)).all()  # noqa: E712
        return [{"id": f.id, "name": f.name, "vendor_id": f.vendor_id} for f in rows]

    # ===== acceptances =====
    @app.post("/api/v1/acceptances")
    def add_acceptance(b: AcceptIn, s: Session = Depends(db),
                       u: User = Depends(require("acceptance.manage"))):
        from datetime import datetime as _dt
        ex = _dt.fromisoformat(b.expires_at) if b.expires_at else None
        row = Acceptance(engagement_id=b.engagement_id, rationale=b.rationale,
                         accepted_by=u.username, expires_at=ex)
        s.add(row); s.flush()
        audit(s, "acceptance.recorded", u.username, {"id": row.id})
        s.commit()
        return {"acceptance_id": row.id, "accepted_by": u.username}

    # ===== contracts (Matt) =====
    @app.post("/api/v1/reassessments")
    def schedule_reassessment(b: ReassessIn, s: Session = Depends(db),
                              u: User = Depends(require("lifecycle.reassess"))):
        row = Reassessment(engagement_id=b.engagement_id, mode=b.mode)
        s.add(row); s.flush()
        audit(s, "reassessment.scheduled", u.username, {"id": row.id, "mode": b.mode})
        notify(s, f"Reassessment scheduled ({b.mode})", "all")
        s.commit()
        return {"reassessment_id": row.id, "mode": b.mode}

    @app.post("/api/v1/reassessments/{rid}/complete")
    def complete_reassessment(rid: int, s: Session = Depends(db),
                              u: User = Depends(require("lifecycle.reassess"))):
        r = s.get(Reassessment, rid)
        if not r:
            raise HTTPException(404, "reassessment not found")
        r.completed = True
        audit(s, "reassessment.completed", u.username, {"id": rid})
        s.commit()
        return {"reassessment_id": rid, "completed": True}

    # ===== corrective action plans (CAP) — modelled as findings =====
    @app.get("/api/v1/cap")
    def cap_board(s: Session = Depends(db),
                  u: User = Depends(require("lifecycle.cap"))):
        rows = s.scalars(select(Finding).where(Finding.status != "closed")).all()
        return {"open_actions": len(rows),
                "by_severity": _count_by(rows, "severity"),
                "items": [{"finding_id": f.id, "title": f.title,
                           "severity": f.severity, "status": f.status} for f in rows]}

    # ===== business impact analysis (BIA) =====
    @app.get("/api/v1/vendors/{vid}/bia")
    def bia(vid: int, s: Session = Depends(db),
            u: User = Depends(require("lifecycle.bia"))):
        _RBAC.assert_object_visible(s, u, 'vendor', vid)
        v = s.get(Vendor, vid)
        if not v:
            raise HTTPException(404, "vendor not found")
        engagements = s.scalars(select(EngagementRow).where(
            EngagementRow.vendor_id == vid)).all()
        return {"vendor_id": vid, "is_critical": v.is_critical,
                "engagement_count": len(engagements),
                "impact": "HIGH" if v.is_critical else
                          "MEDIUM" if len(engagements) > 1 else "LOW"}

    # ===== dashboards =====
    @app.get("/api/v1/dashboard/executive")
    def dash_exec(s: Session = Depends(db),
                  u: User = Depends(require("dashboard.exec"))):
        vendors = s.scalars(select(VendorRecord)).all()
        engs = s.scalars(select(EngagementRecord)).all()
        findings = s.scalars(select(FindingRecord)).all()
        # include any legacy v1 rows so both creation paths are reflected
        v1v = s.scalar(select(func.count()).select_from(Vendor)) or 0
        v1e = s.scalar(select(func.count()).select_from(EngagementRow)) or 0
        v1f = s.scalar(select(func.count()).select_from(Finding)
                       .where(Finding.status != "closed")) or 0
        return {
            "vendors": len(vendors) + v1v,
            "critical_vendors": sum(1 for v in vendors if v.is_critical),
            "engagements": len(engs) + v1e,
            "by_residual": _count_by(engs, "residual_band"),
            "by_decision": _count_by(engs, "status"),
            "open_findings": sum(1 for f in findings if (f.status or "").lower() not in ("closed",)) + v1f,
        }

    @app.get("/api/v1/dashboard/operational")
    def dash_ops(s: Session = Depends(db),
                 u: User = Depends(require("dashboard.ops"))):
        engs = s.scalars(select(EngagementRecord)).all()
        return {"by_stage": _count_by(engs, "status"),
                "by_route": _count_by(engs, "inherent_band"),
                "in_flight": sum(1 for e in engs if (e.status or "") not in ("Terminated", "Exited"))}

    @app.get("/api/v1/dashboard/risk")
    def dash_risk(s: Session = Depends(db),
                  u: User = Depends(require("dashboard.risk"))):
        engs = s.scalars(select(EngagementRow)).all()
        monit = s.scalars(select(Monitoring)).all()
        return {"by_inherent": _count_by(engs, "inherent_band"),
                "by_residual": _count_by(engs, "residual_band"),
                "monitoring_alerts": sum(1 for m in monit if m.status in ("ALERT", "CRITICAL"))}

    # ===== document upload + extraction (feeds Isaac) =====
    @app.patch("/api/v1/vendors/{vid}")
    def update_vendor(vid: int, b: VendorUpdateIn, s: Session = Depends(db),
                      u: User = Depends(require("vendor.edit"))):
        _RBAC.assert_object_visible(s, u, 'vendor', vid)
        v = s.get(Vendor, vid)
        if not v:
            raise HTTPException(404, "vendor not found")
        for f, val in b.model_dump(exclude_none=True).items():
            setattr(v, f, val)
        audit(s, "vendor.updated", u.username, {"vendor_id": vid, "fields": list(b.model_dump(exclude_none=True))})
        s.commit()
        return {"vendor_id": vid, "updated": True}

    @app.delete("/api/v1/vendors/{vid}")
    def delete_vendor(vid: int, s: Session = Depends(db),
                      u: User = Depends(require("vendor.edit"))):
        _RBAC.assert_object_visible(s, u, 'vendor', vid)
        v = s.get(Vendor, vid)
        if not v:
            raise HTTPException(404, "vendor not found")
        # archive semantics: mark, don't hard-delete (preserves audit/history)
        v.ext_ref = (v.ext_ref or "") + "|archived"
        audit(s, "vendor.archived", u.username, {"vendor_id": vid})
        s.commit()
        return {"vendor_id": vid, "archived": True}

    @app.get("/api/v1/vendors/{vid}")
    def get_vendor(vid: int, s: Session = Depends(db),
                   u: User = Depends(require("vendor.view"))):
        _RBAC.assert_object_visible(s, u, 'vendor', vid)
        v = s.get(Vendor, vid)
        if not v:
            raise HTTPException(404, "vendor not found")
        engs = s.scalars(select(EngagementRow).where(EngagementRow.vendor_id == vid)).all()
        return {"vendor_id": v.id, "name": v.name, "industry": v.industry,
                "country": v.country, "contact_email": v.contact_email,
                "tier": v.tier, "is_critical": v.is_critical,
                "critical_reason": v.critical_reason, "critical_by": v.critical_by,
                "engagements": [{"engagement_id": e.id, "title": e.title, "stage": e.stage} for e in engs]}

    # ---- Vendor: list with search/filter (replaces the plain list inline) ----
    @app.get("/api/v1/vendors-search")
    def search_vendors(q: Optional[str] = None, tier: Optional[str] = None,
                       critical: Optional[bool] = None,
                       s: Session = Depends(db), u: User = Depends(require("vendor.view"))):
        # SQL-side filtering + cap — scales to large registers (no full-table
        # Python scan). Archived rows carry '|archived' in ext_ref.
        from sqlalchemy import or_ as _or
        stmt = select(Vendor).where(
            _or(Vendor.ext_ref.is_(None), Vendor.ext_ref.notlike("%|archived%")))
        if q:
            stmt = stmt.where(Vendor.name.ilike(f"%{q.strip()}%"))
        if tier:
            stmt = stmt.where(Vendor.tier == tier)
        if critical is not None:
            stmt = stmt.where(Vendor.is_critical == critical)
        stmt = stmt.order_by(Vendor.name).limit(500)
        return [{"vendor_id": v.id, "name": v.name, "tier": v.tier,
                 "is_critical": v.is_critical} for v in s.scalars(stmt).all()]

    # ---- Vendor: remove critical designation (VRM) ----
    @app.delete("/api/v1/vendors/{vid}/critical")
    def remove_critical(vid: int, s: Session = Depends(db),
                        u: User = Depends(require("vendor.critical"))):
        _RBAC.assert_object_visible(s, u, 'vendor', vid)
        v = s.get(Vendor, vid)
        if not v:
            raise HTTPException(404, "vendor not found")
        v.is_critical = False; v.critical_reason = None; v.critical_by = None
        audit(s, "vendor.critical_removed", u.username, {"vendor_id": vid})
        s.commit()
        return {"vendor_id": vid, "is_critical": False}

    # ---- Engagement: update, list-all, delete ----
    @app.patch("/api/v1/engagements/{eid}")
    def update_engagement(eid: int, b: EngagementUpdateIn, s: Session = Depends(db),
                          u: User = Depends(require("engagement.edit"))):
        _RBAC.assert_object_visible(s, u, 'engagement', eid)
        e = s.get(EngagementRow, eid)
        if not e:
            raise HTTPException(404, "engagement not found")
        for f, val in b.model_dump(exclude_none=True).items():
            setattr(e, f, val)
        audit(s, "engagement.updated", u.username, {"engagement_id": eid})
        s.commit()
        return {"engagement_id": eid, "updated": True}

    @app.get("/api/v1/engagements")
    def list_engagements(stage: Optional[str] = None, vendor_id: Optional[int] = None,
                         s: Session = Depends(db), u: User = Depends(require("engagement.view"))):
        rows = s.scalars(select(EngagementRow)).all()
        out = []
        for e in rows:
            if stage and e.stage != stage:
                continue
            if vendor_id and e.vendor_id != vendor_id:
                continue
            out.append({"engagement_id": e.id, "vendor_id": e.vendor_id, "title": e.title,
                        "stage": e.stage, "route": e.route, "inherent_band": e.inherent_band,
                        "residual_band": e.residual_band, "decision": e.decision})
        return out

    @app.delete("/api/v1/engagements/{eid}")
    def delete_engagement(eid: int, s: Session = Depends(db),
                          u: User = Depends(require("engagement.edit"))):
        _RBAC.assert_object_visible(s, u, 'engagement', eid)
        e = s.get(EngagementRow, eid)
        if not e:
            raise HTTPException(404, "engagement not found")
        e.status = "cancelled"
        audit(s, "engagement.cancelled", u.username, {"engagement_id": eid})
        s.commit()
        return {"engagement_id": eid, "cancelled": True}

    # ---- Finding: update, reopen ----
    @app.post("/api/v1/engagements/{eid}/signoff")
    def signoff(eid: int, b: SignoffIn, s: Session = Depends(db),
                u: User = Depends(require("engagement.review"))):
        _RBAC.assert_object_visible(s, u, 'engagement', eid)
        e = s.get(EngagementRow, eid)
        if not e:
            raise HTTPException(404, "engagement not found")
        e.status = "signed_off" if b.decision == "approved" else "returned"
        audit(s, "engagement.signoff", u.username,
              {"engagement_id": eid, "decision": b.decision, "note": b.note})
        notify(s, f"Engagement #{eid} {b.decision} by VRM", "business")
        s.commit()
        return {"engagement_id": eid, "status": e.status}

    @app.get("/api/v1/review-queue")
    def review_queue(s: Session = Depends(db), u: User = Depends(require("engagement.review"))):
        rows = s.scalars(select(EngagementRow).where(
            EngagementRow.stage == "decision",
            EngagementRow.residual_band.in_(["HIGH", "ELEVATED"]))).all()
        return [{"engagement_id": e.id, "vendor_id": e.vendor_id, "title": e.title,
                 "residual_band": e.residual_band, "decision": e.decision} for e in rows]

    # ---- Auth self-service: change password, profile ----
    @app.post("/api/v1/evidence/{doc_id}/chase")
    def chase_evidence(doc_id: int, s: Session = Depends(db), u: User = Depends(require("lifecycle.evidence"))):
        d = s.get(Document, doc_id)
        if not d:
            raise HTTPException(404, "document not found")
        v = s.get(Vendor, d.vendor_id) if d.vendor_id else None
        to = (v.contact_email if v and v.contact_email else "vendor@example.com")
        from app.features.admin import email_service
        sent = False
        try:
            sent = email_service.send_email(to, f"Evidence renewal required: {d.name}",
                                            "Please submit an updated version of this document.")
        except Exception as _e:
            _obs_swallow('bro_app.py', _e)
        s.add(EmailOutbox(to_addr=to, subject=f"Renewal: {d.name}", body="renewal chase", sent=sent))
        audit(s, "evidence.chased", u.username, {"document_id": doc_id, "to": to})
        notify(s, f"Renewal chased for '{d.name}'", "business")
        s.commit()
        return {"document_id": doc_id, "chased": True, "mode": "smtp" if sent else "simulation"}

    # ---- Contract gap review ----
    @app.post("/api/v1/reassessments/run-due")
    def run_due_reassessments(s: Session = Depends(db), u: User = Depends(require("lifecycle.reassess"))):
        # Tier cadence: Tier1 annual, Tier2 biennial, Tier3 triennial.
        cadence = {"Tier 1": 365, "Tier 2": 730, "Tier 3": 1095}
        created = 0
        for e in s.scalars(select(EngagementRow).where(EngagementRow.stage == "monitor")).all():
            v = s.get(Vendor, e.vendor_id)
            days = cadence.get(v.tier if v else "Tier 3", 1095)
            age = (_dt2.utcnow() - e.created_at).days if e.created_at else 0
            if age >= days:
                s.add(Reassessment(engagement_id=e.id, mode="periodic")); created += 1
        audit(s, "reassessment.cadence_run", u.username, {"created": created})
        s.commit()
        return {"created": created}

    # ---- Reporting: register CSV export, audit export ----
    @app.get("/api/v1/reports/register.csv", response_class=PlainTextResponse)
    def register_csv(s: Session = Depends(db), u: User = Depends(require("reg.report"))):
        buf = _io.StringIO(); w = _csv.writer(buf)
        w.writerow(["vendor_id", "name", "tier", "critical", "engagement_id",
                    "title", "stage", "inherent", "residual", "decision"])
        for v in s.scalars(select(Vendor)).all():
            engs = s.scalars(select(EngagementRow).where(EngagementRow.vendor_id == v.id)).all()
            if not engs:
                w.writerow([v.id, v.name, v.tier, v.is_critical, "", "", "", "", "", ""])
            for e in engs:
                w.writerow([v.id, v.name, v.tier, v.is_critical, e.id, e.title,
                            e.stage, e.inherent_band, e.residual_band, e.decision])
        return buf.getvalue()

    @app.get("/api/v1/audit/export.csv", response_class=PlainTextResponse)
    def audit_export(s: Session = Depends(db), u: User = Depends(require("audit.export")),
                     vendor_id: str = "", engagement_id: str = "", max_rows: int = 100000):
        """Audit export (DB-03): bounded and filterable.

        The full trail is legitimately large, so the export streams in batches rather
        than materialising every row, and can be scoped to one subject."""
        buf = _io.StringIO(); w = _csv.writer(buf)
        w.writerow(["seq", "action", "actor", "vendor_id", "engagement_id", "timestamp", "hash"])
        q = select(AuditLog)
        if vendor_id:
            q = q.where(AuditLog.vendor_id == vendor_id)
        if engagement_id:
            q = q.where(AuditLog.engagement_id == engagement_id)
        cap = min(max(int(max_rows or 100000), 1), 500000)
        written = 0
        for r in s.scalars(q.order_by(AuditLog.seq).limit(cap)).yield_per(1000):
            w.writerow([r.seq, r.action, r.actor, r.vendor_id or "", r.engagement_id or "",
                        r.created_at.isoformat() if r.created_at else "", r.entry_hash])
            written += 1
        if written >= cap:
            w.writerow([f"# truncated at {cap} rows — narrow the filter or raise max_rows"])
        return buf.getvalue()

    # ---- Vendor portal (self-service, scoped) ----
    @app.get("/api/v1/portal/my-status")
    def portal_status(s: Session = Depends(db), u: User = Depends(require("portal.self"))):
        # vendors see a minimal, scoped view (no internal reasoning)
        return {"message": "Vendor portal active",
                "you": u.username,
                "note": "Complete your DDQ and submit evidence via your assigned engagement."}

    # ============================================================
    #  Conversational multi-agent assessment (chat surface)
    # ============================================================
    from app.features.assessment import agents as _A
    from app.features.assessment import agent_engine as _AE
    from app.features.domain.models_feature import AgentLearning, BackgroundInsight

    # ═══ Activity: human actions and agent actions, both immutable ══════════════
    def _activity_visible(s, u, stmt):
        """Scope an audit query to what this user may see.

        Controller and administrator see every user's actions — that is the point of a
        supervisory role. Everyone else sees their own actions, plus system activity on
        records they are permitted to see, so the agent log never becomes a side channel
        around row-level scoping."""
        rk = u.role.key if u.role else None
        if rk in ("admin", "controller"):
            return stmt, "all"
        allowed = _RBAC.scoped_vendor_ids(s, u)
        if allowed is None:                       # assessor / executive: unrestricted
            return stmt, "all"
        return stmt.where(_or2(AuditLog.vendor_id.in_(allowed or ["__none__"]),
                               AuditLog.vendor_id.is_(None))), "scoped"

    def _activity_rows(s, u, kind, *, actor="", action="", vendor_id="",
                       before_seq=0, limit=100):
        from app.features.platform import activity as _ACT
        lim = min(max(int(limit or 100), 1), 500)
        stmt = select(AuditLog)
        stmt, scope = _activity_visible(s, u, stmt)
        rk = u.role.key if u.role else None
        if kind == "user" and rk not in ("admin", "controller"):
            # A user's own log. Supervisors may filter to a person instead.
            stmt = stmt.where(AuditLog.actor == u.username)
        if actor:
            stmt = stmt.where(AuditLog.actor == actor)
        if action:
            stmt = stmt.where(AuditLog.action == action)
        if vendor_id:
            stmt = stmt.where(AuditLog.vendor_id == vendor_id)
        if before_seq:
            stmt = stmt.where(AuditLog.seq < int(before_seq))
        # Over-fetch, then filter by classification: the split between human and agent
        # is derived from the action name rather than stored, so it can be corrected
        # without rewriting history.
        raw = s.scalars(stmt.order_by(AuditLog.seq.desc()).limit(lim * 6)).all()
        want = "human" if kind == "user" else "agent"
        out = []
        for r in raw:
            meta = _ACT.summarise(r.action)
            if meta["kind"] != want:
                continue
            det = {}
            if r.detail:
                try:
                    det = _json2.loads(r.detail)
                except Exception:
                    det = {}
            out.append({"seq": r.seq, "action": r.action, "label": meta["label"],
                        "actor": r.actor, "kind": meta["kind"],
                        "vendor_id": r.vendor_id, "engagement_id": r.engagement_id,
                        "at": r.created_at.isoformat() if r.created_at else None,
                        "hash": r.entry_hash,
                        "detail": {k: v for k, v in list(det.items())[:8]
                                   if not isinstance(v, (dict, list))}})
            if len(out) >= lim:
                break
        return out, scope, (out[-1]["seq"] if out and len(out) >= lim else None)

    @app.get("/api/v1/activity/user")
    def activity_user(s: Session = Depends(db),
                      u: User = Depends(require("engagement.view")),
                      actor: str = "", action: str = "", vendor_id: str = "",
                      before_seq: int = 0, limit: int = 100):
        """Actions performed by people. Every user sees their own; a controller or
        administrator sees everyone's. Entries are drawn from the hash-chained audit
        trail and are immutable — this is a view over the record, not a second copy."""
        rows, scope, nxt = _activity_rows(s, u, "user", actor=actor, action=action,
                                          vendor_id=vendor_id, before_seq=before_seq,
                                          limit=limit)
        return {"entries": rows, "count": len(rows), "scope": scope,
                "viewer": u.username, "next_before_seq": nxt,
                "immutable": True,
                "note": ("Sourced from the tamper-evident audit chain. Verify integrity "
                         "at /api/v1/audit/verify.")}

    @app.get("/api/v1/activity/agent")
    def activity_agent(s: Session = Depends(db),
                       u: User = Depends(require("engagement.view")),
                       action: str = "", vendor_id: str = "",
                       before_seq: int = 0, limit: int = 100):
        """Actions performed by the system: automated record chains, scheduled sweeps,
        screening outcomes and AI activity — scoped to the records this user may see.

        These remain attributed to the human who invoked them, because the model is
        never the accountable party. The classification says what performed the action,
        not who is answerable for it."""
        rows, scope, nxt = _activity_rows(s, u, "agent", action=action,
                                          vendor_id=vendor_id, before_seq=before_seq,
                                          limit=limit)
        return {"entries": rows, "count": len(rows), "scope": scope,
                "viewer": u.username, "next_before_seq": nxt, "immutable": True,
                "note": ("Agent actions stay attributed to the person who invoked them "
                         "(SOP IAM-4). This log says what acted, not who is answerable.")}

    @app.get("/api/v1/activity/summary")
    def activity_summary(s: Session = Depends(db),
                         u: User = Depends(require("engagement.view"))):
        """Counts for the activity page tabs."""
        from app.features.platform import activity as _ACT
        stmt, scope = _activity_visible(s, u, select(AuditLog))
        raw = s.scalars(stmt.order_by(AuditLog.seq.desc()).limit(2000)).all()
        h = a = 0
        actors = {}
        for r in raw:
            if _ACT.classify(r.action) == "agent":
                a += 1
            else:
                h += 1
                actors[r.actor] = actors.get(r.actor, 0) + 1
        return {"human_actions": h, "agent_actions": a, "scope": scope,
                "top_actors": sorted(actors.items(), key=lambda x: -x[1])[:8],
                "can_see_all_users": (u.role.key if u.role else "") in
                                     ("admin", "controller")}

    @app.get("/api/v1/platform/run-budgets")
    def platform_run_budgets(u: User = Depends(require("admin.config"))):
        """In-flight run budgets (AI-05). A run that is 90% through its token budget is
        the signal that something is regenerating rather than converging."""
        from app.features.platform import budget as _BUD
        runs = _BUD.active_runs()
        return {"active": runs, "count": len(runs),
                "near_limit": [r["run_id"] for r in runs
                               if (r.get("tokens_pct") or 0) > 80
                               or (r.get("cost_pct") or 0) > 80]}

    @app.get("/api/v1/platform/reliability")
    def platform_reliability(u: User = Depends(require("admin.config"))):
        """Circuit-breaker states (APP-03). An open circuit is a dependency the
        platform has stopped calling; operators need to see that without reading logs."""
        from app.features.platform import reliability as _REL
        st = _REL.breaker_states()
        return {"breakers": st,
                "open": [b["name"] for b in st if b["state"] == "open"],
                "healthy": all(b["state"] == "closed" for b in st)}

    @app.get("/api/v1/ai/cache-metrics")
    def ai_cache_metrics(s: Session = Depends(db),
                         u: User = Depends(require("admin.aikeys"))):
        """Prompt-cache effectiveness per domain (AI-03).

        Cache hit rate is a first-class production metric: a cache that never hits is
        a standing surcharge on input tokens rather than a saving. Reads bill at a
        fraction of base input and writes at a premium, so the write:read ratio is
        what says whether caching is paying for itself."""
        from sqlalchemy import text as _t
        try:
            rows = s.execute(_t(
                "SELECT domain, COUNT(*), COALESCE(SUM(input_tokens),0), "
                "COALESCE(SUM(cache_read_tokens),0), COALESCE(SUM(cache_write_tokens),0) "
                "FROM ai_call_log WHERE success=1 GROUP BY domain "
                "ORDER BY COUNT(*) DESC")).fetchall()
        except Exception:
            return {"available": False, "reason": "ledger not initialised"}
        out, t_cr, t_cw, t_in = [], 0, 0, 0
        for d, calls, inp, cr, cw in rows:
            total = (cr or 0) + (cw or 0) + (inp or 0)
            out.append({"domain": d or "-", "calls": calls, "input_tokens": inp,
                        "cache_read_tokens": cr, "cache_write_tokens": cw,
                        "hit_rate": round((cr / total) * 100, 1) if total else 0.0,
                        "write_read_ratio": (round(cw / cr, 2) if cr else None)})
            t_cr += cr or 0; t_cw += cw or 0; t_in += inp or 0
        grand = t_cr + t_cw + t_in
        return {"available": True, "by_domain": out,
                "overall_hit_rate": round((t_cr / grand) * 100, 1) if grand else 0.0,
                "overall_write_read_ratio": (round(t_cw / t_cr, 2) if t_cr else None),
                "note": ("Write:read below 1.0 means caching is paying for itself. A high "
                         "write:read with a low hit rate means the cache breakpoint is "
                         "sitting on volatile content.")}

    @app.get("/api/v1/ai/status")
    def ai_status(u: User = Depends(require("admin.integrations"))):
        from app.agents import llm_config
        st = dict(llm_config.status())
        st["claude_key_present"] = bool(_os.environ.get("ANTHROPIC_API_KEY"))
        st["openai_key_present"] = bool(_os.environ.get("OPENAI_API_KEY"))
        st["grok_key_present"] = bool(_os.environ.get("XAI_API_KEY"))
        st["manus_key_present"] = bool(_os.environ.get("MANUS_API_KEY"))
        st["nvidia_key_present"] = bool(_os.environ.get("NVIDIA_API_KEY"))
        st["providers"] = llm_config.provider_info()
        st["secrets_encrypted"] = SEC.secret_key_set()
        st["production_mode"] = SEC.is_production()
        return st

    @app.get("/api/v1/ai/models")
    def ai_models(u: User = Depends(require("admin.integrations"))):
        """List the models the configured key can actually access, so the admin
        can pick a valid one instead of guessing."""
        from app.agents import llm_config
        st = llm_config.status()
        prov = st.get("provider")
        if not st.get("live_ready"):
            return {"available": False, "models": [],
                    "error": ("no provider key configured" if not prov
                              else f"{prov} SDK not installed")}
        try:
            ids = []
            if prov == "claude":
                import anthropic
                key = _os.environ.get("ANTHROPIC_API_KEY") or _os.environ.get("BRO_LLM_KEY")
                for m in anthropic.Anthropic(api_key=key).models.list(limit=50).data:
                    ids.append(getattr(m, "id", None))
            elif prov == "openai":
                import openai
                cl = openai.OpenAI(api_key=_os.environ.get("OPENAI_API_KEY"))
                ids = [getattr(m, "id", None) for m in cl.models.list().data
                       if "gpt" in (getattr(m, "id", "") or "")]
            ids = [i for i in ids if i]
            return {"available": True, "provider": prov, "models": ids,
                    "current": st.get("model")}
        except Exception as e:
            return {"available": False, "models": [],
                    "error": f"{type(e).__name__}: {str(e)[:200]}"}

    @app.post("/api/v1/ai/test")
    def ai_test(u: User = Depends(require("admin.integrations"))):
        """Live diagnostic: makes a real call (and a web-search call) so the
        admin can see exactly why AI isn't working, rather than silent fallback."""
        from app.agents import llm_config
        st = llm_config.status()
        if not st.get("live_ready"):
            return {"live_ready": False, "basic_ok": False, "web_ok": False,
                    "error": ("AI is not live: " + ("no provider key configured" if not st.get("provider")
                              else f"{st.get('provider')} SDK not installed"))}
        basic = llm_config.complete("You are a connectivity probe.",
                                    "Reply with exactly: OK", domain="general")
        basic_err = llm_config.last_error()
        web = llm_config.complete("You are a web-search probe with live search.",
                                  "Search the web and reply with one current headline.",
                                  domain="general", web_search=True)
        web_err = llm_config.last_error()
        return {"live_ready": True, "provider": st.get("provider"), "model": st.get("model"),
                "basic_ok": bool(basic), "basic_reply": (basic or "")[:120], "basic_error": basic_err,
                "web_ok": bool(web), "web_reply": (web or "")[:160], "web_error": web_err}

    _AI_ENV = {"claude": "ANTHROPIC_API_KEY", "openai": "OPENAI_API_KEY",
               "grok": "XAI_API_KEY", "manus": "MANUS_API_KEY",
               "nvidia": "NVIDIA_API_KEY"}

    def _persist_ai_keys(s: Session, updates: dict):
        """Persist provider keys in the system config so they survive restart."""
        from app.features.domain import config_store as CFG
        store = CFG.get_json(s, "ai_provider_keys", {}) or {}
        store.update({k: SEC.encrypt_value(v) for k, v in updates.items() if v})
        CFG.upsert_json(s, "ai_provider_keys", store, updated_by="admin", category="ai")
        s.commit()

    @app.post("/api/v1/ai/key")
    def ai_set_key(b: AiKeyIn, s: Session = Depends(db),
                   u: User = Depends(require("admin.aikeys"))):
        """Set a live LLM provider key. The key is applied to the running process and
        persisted in the system configuration so it survives restart."""
        from app.agents import llm_config
        provider = (b.provider or "claude").lower().strip()
        key = (b.api_key or "").strip()
        if provider not in _AI_ENV:
            raise HTTPException(400, "provider must be one of: claude, openai, grok, manus")
        if key:
            _os.environ[_AI_ENV[provider]] = key
            _persist_ai_keys(s, {provider: key})
        # all four are now wired as live providers (grok/manus via OpenAI-compatible base)
        _os.environ["BRO_LLM_PROVIDER"] = provider
        if (b.model or "").strip():
            _os.environ["BRO_LLM_MODEL"] = b.model.strip()
        try:
            llm_config._adapter.cache_clear()
            llm_config._LAST_ERROR = None
        except Exception as _e:
            _obs_swallow('bro_app.py', _e)
        audit(s, "ai.key_set", u.username, {"provider": provider})
        return ai_status(u)

    @app.post("/api/v1/ai/keys")
    def ai_set_keys(body: dict = Body(default={}), s: Session = Depends(db),
                    u: User = Depends(require("admin.aikeys"))):
        """Bulk-set provider keys from the AI integration panel.
        Body: {anthropic, openai, grok, manus} — only non-empty values are applied."""
        from app.agents import llm_config
        alias = {"anthropic": "claude", "claude": "claude", "openai": "openai",
                 "grok": "grok", "xai": "grok", "manus": "manus",
                 "nvidia": "nvidia", "nim": "nvidia"}
        updates = {}
        for k, v in (body or {}).items():
            prov = alias.get(k.lower())
            if prov and (v or "").strip():
                _os.environ[_AI_ENV[prov]] = v.strip()
                updates[prov] = v.strip()
        if updates:
            _persist_ai_keys(s, updates)
            audit(s, "ai.keys_set", u.username, {"providers": sorted(updates)})
            try:
                llm_config._adapter.cache_clear()
                llm_config._LAST_ERROR = None
            except Exception as _e:
                _obs_swallow('bro_app.py', _e)
        return ai_status(u)

    @app.post("/api/v1/ai/test/{provider}")
    def ai_test_provider(provider: str, u: User = Depends(require("admin.integrations"))):
        """Probe a single provider's key/endpoint with a tiny live call (independent of
        the currently-selected provider), so the admin can confirm a key before relying on it."""
        from app.agents import llm_config
        return llm_config.test_provider(provider)

    @app.post("/api/v1/ai/provider")
    def ai_set_provider(body: dict = Body(default={}), s: Session = Depends(db),
                        u: User = Depends(require("admin.aikeys"))):
        """Select the active live inference provider (claude / openai / grok / manus)
        and persist the choice so it survives restart."""
        from app.agents import llm_config
        from app.features.domain import config_store as CFG
        prov = (body.get("provider") or "").lower().strip()
        if prov not in llm_config.all_provider_ids():
            raise HTTPException(400, f"unknown provider: {prov}")
        _os.environ["BRO_LLM_PROVIDER"] = prov
        CFG.upsert_json(s, "ai_active_provider", {"provider": prov},
                        updated_by="admin", category="ai")
        try:
            llm_config._adapter.cache_clear()
            llm_config._LAST_ERROR = None
        except Exception as _e:
            _obs_swallow('bro_app.py', _e)
        audit(s, "ai.provider_selected", u.username, {"provider": prov})
        s.commit()
        return ai_status(u)

    # ===== v4.2: interconnected ecosystem =====
    @app.get("/api/v1/search")
    def global_search(q: str = "", s: Session = Depends(db),
                      u: User = Depends(require("vendor.view"))):
        ql = (q or "").strip()
        if len(ql) < 2:
            return {"q": q, "results": []}
        like = f"%{ql}%"
        res = []
        from app.features.domain.registry_models import EngagementRecord, IncidentRecord
        from sqlalchemy import or_ as _or

        # Vendors — filtered and capped in SQL, not in Python.
        for v in s.scalars(
            select(VendorRecord).where(_or(
                VendorRecord.legal_name.ilike(like),
                VendorRecord.vendor_id.ilike(like))).limit(8)
        ).all():
            res.append({"kind": "vendor", "id": v.vendor_id, "title": v.legal_name,
                        "sub": f"{v.vendor_id} · {v.tier}" + (" · CRITICAL" if v.is_critical else "")})

        # Engagements
        for e in s.scalars(
            select(EngagementRecord).where(_or(
                EngagementRecord.title.ilike(like),
                EngagementRecord.engagement_id.ilike(like))).limit(6)
        ).all():
            res.append({"kind": "engagement", "id": e.engagement_id, "title": e.title,
                        "sub": f"{e.engagement_id} · {e.vendor_id} · residual {e.residual_band or '—'}",
                        "vendor_id": e.vendor_id})

        # Incidents
        for i in s.scalars(
            select(IncidentRecord).where(_or(
                IncidentRecord.incident_id.ilike(like),
                IncidentRecord.incident_type.ilike(like),
                IncidentRecord.vendor_name.ilike(like),
                IncidentRecord.impact_description.ilike(like))).limit(4)
        ).all():
            res.append({"kind": "incident", "id": i.incident_id,
                        "title": f"{i.incident_type or 'Incident'} · {i.vendor_name or i.vendor_id or ''}",
                        "sub": f"{i.incident_id} · {i.severity or '—'} · {i.status or '—'}",
                        "vendor_id": i.vendor_id})

        # Contracts (optional module)
        try:
            from .features.contract_models import ContractRecord
            for c in s.scalars(
                select(ContractRecord).where(ContractRecord.title.ilike(like)).limit(2)
            ).all():
                res.append({"kind": "contract", "id": c.id, "title": c.title,
                            "sub": f"contract · {c.vendor_id or ''}", "vendor_id": c.vendor_id})
        except Exception as _e:
            _obs_swallow('bro_app.py', _e)
        return {"q": q, "results": res[:20]}

    @app.get("/api/v1/connections/mcp")
    def mcp_list(s: Session = Depends(db), u: User = Depends(require("admin.integrations"))):
        from app.features.domain import config_store as CFG
        return {"connections": CFG.get_json(s, "mcp_connections", []) or []}

    @app.post("/api/v1/connections/mcp")
    def mcp_add(body: dict = Body(default={}), s: Session = Depends(db),
                u: User = Depends(require("admin.integrations"))):
        from app.features.domain import config_store as CFG
        name = (body.get("name") or "").strip(); url = (body.get("url") or "").strip()
        if not (name and url):
            raise HTTPException(400, "name and url are required")
        lst = [c for c in (CFG.get_json(s, "mcp_connections", []) or []) if c.get("name") != name]
        lst.append({"name": name, "url": url, "transport": body.get("transport") or "sse",
                    "auth": body.get("auth") or "none", "status": "configured (untested)",
                    "added": __import__("datetime").date.today().isoformat()})
        CFG.upsert_json(s, "mcp_connections", lst, updated_by=u.username, category="connections")
        audit(s, "connections.mcp_add", u.username, {"name": name}); s.commit()
        return {"connections": lst}

    @app.post("/api/v1/connections/mcp/{name}/delete")
    def mcp_del(name: str, s: Session = Depends(db),
                u: User = Depends(require("admin.integrations"))):
        from app.features.domain import config_store as CFG
        lst = [c for c in (CFG.get_json(s, "mcp_connections", []) or []) if c.get("name") != name]
        CFG.upsert_json(s, "mcp_connections", lst, updated_by=u.username, category="connections")
        s.commit(); return {"connections": lst}

    @app.get("/api/v1/schedules")
    def schedules_list(s: Session = Depends(db), u: User = Depends(require("vendor.view"))):
        from app.features.domain import config_store as CFG
        from sqlalchemy import text as _t
        last_run = None
        try:
            last_run = s.execute(_t("SELECT MAX(started_at) FROM monitoring_runs")).scalar()
        except Exception as _e:
            _obs_swallow('bro_app.py', _e)
        defaults = [
            {"id": "monitoring_sweep", "label": "Continuous monitoring sweep", "what": "Re-screens the portfolio: alerts, thresholds, stage clocks.", "engine": "in-process scheduler / Render cron", "cadence_hours": _monitor_interval()},
            {"id": "sanctions_screen", "label": "Sanctions & AML re-screening", "what": "Re-screens vendors and persons against sanctions lists.", "engine": "on-demand today", "cadence_hours": 24},
            {"id": "fdd_refresh", "label": "Financial health refresh (Vera)", "what": "Refreshes RapidRatings-style FDD signals on monitored vendors.", "engine": "on-demand today", "cadence_hours": 168},
            {"id": "cert_expiry_scan", "label": "Certification expiry scan", "what": "Flags ISO/SOC certificates entering their expiry window.", "engine": "on-demand today", "cadence_hours": 168},
            {"id": "exit_trigger_scan", "label": "Exit trigger scan", "what": "Evaluates exit-trigger conditions across critical vendors.", "engine": "on-demand today", "cadence_hours": 168},
            {"id": "contract_expiry_scan", "label": "Contract renewal window scan", "what": "Surfaces contracts entering notice/renewal windows.", "engine": "on-demand today", "cadence_hours": 168},
        ]
        cfg = CFG.get_json(s, "schedules", {}) or {}
        out = []
        for d in defaults:
            o = cfg.get(d["id"], {})
            d["enabled"] = bool(o.get("enabled", d["id"] == "monitoring_sweep" and _os.environ.get("BRO_SCHEDULER_ENABLED") == "1"))
            d["cadence_hours"] = int(o.get("cadence_hours", d["cadence_hours"]))
            out.append(d)
        return {"scheduler_running": _os.environ.get("BRO_SCHEDULER_ENABLED") == "1",
                "monitoring_last_run": last_run, "schedules": out}

    @app.post("/api/v1/schedules/{sid}")
    def schedules_set(sid: str, body: dict = Body(default={}), s: Session = Depends(db),
                      u: User = Depends(require("admin.integrations"))):
        from app.features.domain import config_store as CFG
        from app.features.admin import notifications as NOTIF
        cfg = CFG.get_json(s, "schedules", {}) or {}
        e = cfg.setdefault(sid, {})
        if "enabled" in body:
            e["enabled"] = bool(body["enabled"])
        if body.get("cadence_hours"):
            e["cadence_hours"] = max(1, int(body["cadence_hours"]))
        CFG.upsert_json(s, "schedules", cfg, updated_by=u.username, category="schedules")
        NOTIF.emit(s, "schedule.changed", f"Schedule '{sid}' updated",
                   f"enabled={e.get('enabled')} cadence={e.get('cadence_hours')}h by {u.username}")
        audit(s, "schedules.update", u.username, {"id": sid, **e}); s.commit()
        return {"id": sid, **e}

    @app.post("/api/v1/ai/dump-to-draft")
    async def dump_to_draft(files: list[UploadFile] = File(...), fields: str = Form(...),
                            context: str = Form(default=""), s: Session = Depends(db),
                            u: User = Depends(require("vendor.view"))):
        import json as _json
        from app.features.assessment import draftfill as DF
        try:
            field_list = _json.loads(fields); assert isinstance(field_list, list)
        except Exception:
            raise HTTPException(400, "fields must be a JSON list of {id,label}")
        texts = []
        for f in files[:6]:
            data = await f.read()
            _err = SEC.upload_check(f.filename or "doc", data)
            if _err:
                raise HTTPException(415, _err)
            t = DF.extract_text(f.filename or "", data)
            if t.strip():
                texts.append(t)
        corpus = chr(10).join(texts)
        if not corpus.strip():
            raise HTTPException(422, "no readable text found in the uploaded document(s)")
        values, engine = None, "rules"
        ai = DF.ai_fill(s, field_list, corpus, context)
        if ai:
            ids = {f.get("id") for f in field_list}
            values = {k: v for k, v in ai.items() if k in ids and v not in (None, "")}
            engine = "ai"
        if not values:
            values = DF.heuristic_fill(field_list, corpus); engine = "rules"
        audit(s, "ai.dump_to_draft", u.username, {"fields": len(field_list), "filled": len(values), "engine": engine})
        s.commit()
        return {"engine": engine, "values": values, "filled": len(values), "of": len(field_list)}

    @app.get("/api/v1/ai/ledger")
    def ai_ledger_view(s: Session = Depends(db),
                       u: User = Depends(require("admin.integrations"))):
        """AI call ledger (metadata only — no prompt/response content is ever stored)."""
        from app.features.assessment import ai_ledger as AL
        return {"recent": AL.recent(s, 25), "today_count": AL.today_count(s),
                "daily_budget": AL.get_budget(s)}

    @app.post("/api/v1/ai/budget")
    def ai_set_budget(body: dict = Body(default={}), s: Session = Depends(db),
                      u: User = Depends(require("admin.aikeys"))):
        from app.features.assessment import ai_ledger as AL
        v = body.get("daily_calls")
        AL.set_budget(s, int(v) if v is not None and str(v) != "" else None)
        audit(s, "ai.budget_set", u.username, {"daily_calls": v})
        return {"daily_budget": AL.get_budget(s)}

    @app.post("/api/v1/ai/custom-provider")
    def ai_add_custom_provider(body: dict = Body(default={}), s: Session = Depends(db),
                               u: User = Depends(require("admin.aikeys"))):
        """Register ANY new OpenAI-compatible LLM service (label, base URL, model, key)."""
        from app.agents import llm_config
        from app.features.domain import config_store as CFG
        label = (body.get("label") or "").strip()
        base_url = (body.get("base_url") or "").strip()
        model = (body.get("model") or "").strip()
        api_key = (body.get("api_key") or "").strip()
        pid = llm_config._slug(body.get("id") or label)
        if not (label and base_url and model):
            raise HTTPException(400, "label, base_url and model are required")
        if pid in llm_config._BUILTIN_ORDER:
            raise HTTPException(400, f"id '{pid}' collides with a built-in provider")
        llm_config.register_provider(pid, label, base_url, model)
        lst = [c for c in (CFG.get_json(s, "ai_custom_providers", []) or []) if c.get("id") != pid]
        lst.append({"id": pid, "label": label, "base_url": base_url, "model": model})
        CFG.upsert_json(s, "ai_custom_providers", lst, updated_by="admin", category="ai")
        if api_key:
            _os.environ[llm_config._PROVIDER_KEY_ENV[pid]] = api_key
            _persist_ai_keys(s, {pid: api_key})
        audit(s, "ai.custom_provider_add", u.username, {"id": pid, "base_url": base_url})
        s.commit()
        return ai_status(u)

    @app.post("/api/v1/ai/custom-provider/{pid}/delete")
    def ai_del_custom_provider(pid: str, s: Session = Depends(db),
                               u: User = Depends(require("admin.aikeys"))):
        from app.agents import llm_config
        from app.features.domain import config_store as CFG
        pid = llm_config._slug(pid)
        llm_config.unregister_provider(pid)
        lst = [c for c in (CFG.get_json(s, "ai_custom_providers", []) or []) if c.get("id") != pid]
        CFG.upsert_json(s, "ai_custom_providers", lst, updated_by="admin", category="ai")
        store = CFG.get_json(s, "ai_provider_keys", {}) or {}
        store.pop(pid, None)
        CFG.upsert_json(s, "ai_provider_keys", store, updated_by="admin", category="ai")
        if _os.environ.get("BRO_LLM_PROVIDER") == pid:
            _os.environ.pop("BRO_LLM_PROVIDER", None)
        audit(s, "ai.custom_provider_del", u.username, {"id": pid})
        s.commit()
        return ai_status(u)

    @app.get("/api/v1/ai/prompts")
    def ai_list_prompts(s: Session = Depends(db),
                        u: User = Depends(require("admin.aikeys"))):
        """All editable AI-feature prompts, with current text (override or default)."""
        return {"prompts": PROMPTS.listing(s), "groups": PROMPTS.GROUP_ORDER}

    @app.post("/api/v1/ai/prompts/{key}")
    def ai_set_prompt(key: str, body: dict = Body(...), s: Session = Depends(db),
                      u: User = Depends(require("admin.aikeys"))):
        text = (body.get("text") or "").strip()
        if not text:
            raise HTTPException(400, "text required")
        # AI-04: a change note is required. "Why did this prompt change?" is the first
        # question asked when an assessment is challenged, and it cannot be answered
        # retrospectively.
        note = (body.get("note") or "").strip()
        if not note:
            raise HTTPException(400, "note required: describe why this prompt is changing")
        try:
            PROMPTS.set_prompt(s, key, text, by=u.username, note=note)
        except KeyError:
            raise HTTPException(404, f"unknown prompt key: {key}")
        ver = PROMPTS.prompt_version(s, key)
        audit(s, "ai.prompt_set", u.username,
              {"key": key, "len": len(text), "version": ver["version"], "note": note})
        s.commit()
        return {"ok": True, "key": key, "overridden": True, "current": text,
                "version": ver["version"],
                "reminder": ("Run tools/release_gate.sh before relying on this change — "
                             "a prompt edit alters assessment behaviour like a code change.")}

    @app.get("/api/v1/ai/prompts/{key}/version")
    def ai_prompt_version(key: str, s: Session = Depends(db),
                          u: User = Depends(require("admin.aikeys"))):
        """Version, owner and changelog for a prompt (AI-04)."""
        return PROMPTS.prompt_version(s, key)

    @app.post("/api/v1/ai/prompts/{key}/reset")
    def ai_reset_prompt(key: str, s: Session = Depends(db),
                        u: User = Depends(require("admin.aikeys"))):
        if key not in PROMPTS.defaults():
            raise HTTPException(404, f"unknown prompt key: {key}")
        PROMPTS.reset_prompt(s, key, by=u.username)
        audit(s, "ai.prompt_reset", u.username, {"key": key})
        s.commit()
        return {"ok": True, "key": key, "overridden": False,
                "current": PROMPTS.defaults()[key]["default"]}

    @app.post("/api/v1/ai/key/clear")
    def ai_clear_key(u: User = Depends(require("admin.aikeys"))):
        """Remove any in-session key and fall back to the deterministic engines."""
        from app.agents import llm_config
        for k in ("ANTHROPIC_API_KEY", "OPENAI_API_KEY", "BRO_LLM_KEY",
                  "BRO_LLM_PROVIDER", "BRO_LLM_MODEL"):
            _os.environ.pop(k, None)
        try:
            llm_config._adapter.cache_clear()
            llm_config._LAST_ERROR = None
        except Exception as _e:
            _obs_swallow('bro_app.py', _e)
        return llm_config.status()

    @app.get("/api/v2/industries")
    def list_industries(s: Session = Depends(db), u: User = Depends(require("vendor.view"))):
        return [{"industry_id": i.industry_id, "sic_code": i.sic_code, "division": i.division}
                for i in s.scalars(select(IndustryMaster).order_by(IndustryMaster.sic_code)).all()]

    @app.get("/api/v2/material-groups")
    def list_material_groups(s: Session = Depends(db), u: User = Depends(require("engagement.view"))):
        return [{"material_group_id": m.material_group_id, "unspsc_code": m.unspsc_code}
                for m in s.scalars(select(MaterialGroupMaster).order_by(MaterialGroupMaster.unspsc_code)).all()]

    # ---- vendors (exhaustive) ----
    @app.post("/api/v2/vendors")
    def v2_create_vendor(b: V2VendorIn, request: Request, s: Session = Depends(db),
                         u: User = Depends(require("vendor.edit"))):
        """APP-02: honours an Idempotency-Key header. A retried creation returns the
        first result instead of minting a second supplier — the register is the system
        of record and duplicates in it reach the regulator's evidence pack."""
        from app.features.platform import reliability as _REL

        def _create():
            return _v2_create_vendor_inner(b, s, u)

        try:
            return _REL.idempotent(
                s, key=request.headers.get("Idempotency-Key"), actor=u.username,
                route="POST /api/v2/vendors", body=b.model_dump(), produce=_create)
        except _REL.IdempotencyConflict as _e:
            raise HTTPException(409, str(_e))

    def _v2_create_vendor_inner(b, s, u):
        v = RS.create_vendor(s, legal_name=b.legal_name, created_via=b.created_via or "button",
                             group_id=b.group_id, parent=b.parent_company,
                             industries=b.industries or [], tier=b.tier or "Tier 3",
                             trading_name=b.trading_name, registration_number=b.registration_number,
                             hq_country=b.hq_country, website=b.website,
                             listing_status=b.listing_status, procurement_ref=b.procurement_ref)
        audit(s, "v2.vendor_created", u.username,
              {"vendor_id": v.vendor_id, "group_id": v.group_id, "via": v.created_via})
        s.commit()
        return {"vendor_id": v.vendor_id, "group_id": v.group_id}

    @app.post("/api/v2/vendors/import")
    async def v2_import_vendors(file: UploadFile = File(...), mode: str = Form(default="preview"),
                                s: Session = Depends(db), u: User = Depends(require("vendor.edit"))):
        """Bulk CSV vendor import. Columns: legal_name (required), tier, hq_country,
        is_critical. mode=preview validates only; mode=commit creates the vendors."""
        import csv as _csv
        import io as _io
        raw = await file.read()
        _err = SEC.upload_check(file.filename or "vendors.csv", raw)
        if _err:
            raise HTTPException(415, _err)
        try:
            textd = raw.decode("utf-8-sig")
        except Exception:
            raise HTTPException(415, "file must be UTF-8 CSV")
        rows = list(_csv.DictReader(_io.StringIO(textd)))
        if not rows:
            raise HTTPException(400, "no data rows found (expected a header incl. legal_name)")
        valid, errors, seen = [], [], set()
        existing = {(v.legal_name or "").strip().lower()
                    for v in s.scalars(select(VendorRecord)).all()}
        for i, r in enumerate(rows, start=2):
            name = (r.get("legal_name") or r.get("name") or "").strip()
            if not name:
                errors.append({"row": i, "error": "legal_name is required"})
                continue
            key = name.lower()
            if key in existing:
                errors.append({"row": i, "error": f"'{name}' already exists in the register"})
                continue
            if key in seen:
                errors.append({"row": i, "error": f"'{name}' duplicated in file"})
                continue
            seen.add(key)
            tier = (r.get("tier") or "Tier 3").strip()
            valid.append({"legal_name": name,
                          "tier": tier if tier in ("Tier 1", "Tier 2", "Tier 3") else "Tier 3",
                          "hq_country": (r.get("hq_country") or "").strip() or None,
                          "is_critical": str(r.get("is_critical", "")).strip().lower()
                          in ("1", "true", "yes", "y")})
        created = []
        if mode == "commit":
            for v in valid:
                rec = RS.create_vendor(s, legal_name=v["legal_name"], created_via="csv_import",
                                       tier=v["tier"], hq_country=v["hq_country"], industries=[])
                if v["is_critical"]:
                    rec.is_critical = True
                created.append(rec.vendor_id)
            audit(s, "v2.vendors_imported", u.username, {"count": len(created)})
            from app.features.admin import notifications as _NOTIF
            _NOTIF.emit(s, "import.completed", f"CSV import: {len(created)} vendor(s) created", f"by {u.username}", link="vendors")
            for _vid in created:
                _NOTIF.emit(s, "vendor.created", f"Vendor created: {_vid}", "via CSV import", link="vendors")
            s.commit()
        return {"mode": mode, "total_rows": len(rows), "valid": len(valid),
                "errors": errors[:50], "created": created}

    @app.get("/api/v2/vendors")
    def v2_list_vendors(group_id: Optional[str] = None, slim: bool = False,
                        s: Session = Depends(db),
                        u: User = Depends(require("vendor.view"))):
        stmt = select(VendorRecord)
        if group_id:
            stmt = stmt.where(VendorRecord.group_id == group_id)
        # row-level isolation: supplier sees only their vendor; buyer only owned
        _allowed = _RBAC.scoped_vendor_ids(s, u)
        if _allowed is not None:
            stmt = stmt.where(VendorRecord.vendor_id.in_(_allowed or ["__none__"]))
        vendors = s.scalars(stmt).all()
        from app.features.domain import watchlist_service as _WL
        _wl_ids = _WL.watchlisted_vendor_ids(s)
        if slim:
            # SPA list view uses only these fields; skip the industries join entirely.
            return [{"vendor_id": v.vendor_id, "legal_name": v.legal_name, "tier": v.tier,
                     "status": v.status, "is_critical": v.is_critical,
                     "watchlisted": v.vendor_id in _wl_ids} for v in vendors]
        # Batch all industry tags in ONE query instead of N (was an N+1 per vendor).
        vids = [v.vendor_id for v in vendors]
        ind_map: dict = {}
        if vids:
            for vi in s.scalars(select(VendorIndustry).where(
                    VendorIndustry.vendor_id.in_(vids))).all():
                ind_map.setdefault(vi.vendor_id, []).append(vi.industry_id)
        return [{"vendor_id": v.vendor_id, "group_id": v.group_id,
                 "legal_name": v.legal_name, "tier": v.tier, "status": v.status,
                 "is_critical": v.is_critical, "industries": ind_map.get(v.vendor_id, []),
                 "watchlisted": v.vendor_id in _wl_ids,
                 "created_via": v.created_via} for v in vendors]

    @app.get("/api/v2/vendors/{vid}")
    def v2_get_vendor(vid: str, s: Session = Depends(db), u: User = Depends(require("vendor.view"))):
        _RBAC.assert_object_visible(s, u, 'vendor', vid)
        v = s.scalars(select(VendorRecord).where(VendorRecord.vendor_id == vid)).first()
        if not v:
            raise HTTPException(404, "vendor not found")
        if not _RBAC.can_see_vendor(s, u, vid):
            raise HTTPException(403, "not authorised for this vendor")
        inds = [vi.industry_id for vi in s.scalars(select(VendorIndustry).where(
            VendorIndustry.vendor_id == vid)).all()]
        contacts = [{"id": c.id, "is_primary": c.is_primary, "name": c.name, "email": c.email,
                     "phone": f"{c.phone_country_code or ''} {c.phone_number or ''}".strip(),
                     "designation": c.designation, "country": c.country,
                     "mailing_address": c.mailing_address}
                    for c in s.scalars(select(ContactRecord).where(
                        ContactRecord.owner_type == "vendor", ContactRecord.owner_id == vid)).all()]
        engs = [{"engagement_id": e.engagement_id, "title": e.title, "status": e.status}
                for e in s.scalars(select(EngagementRecord).where(
                    EngagementRecord.vendor_id == vid)).all()]
        return {"vendor_id": v.vendor_id, "group_id": v.group_id, "legal_name": v.legal_name,
                "trading_name": v.trading_name, "tier": v.tier, "status": v.status,
                "hq_country": v.hq_country, "website": v.website,
                "listing_status": v.listing_status, "is_critical": v.is_critical,
                "industries": inds, "contacts": contacts, "engagements": engs,
                "fourth_party_id": v.fourth_party_id}

    @app.post("/api/v2/vendors/{vid}/group")
    def v2_override_group(vid: str, b: GroupOverrideIn, s: Session = Depends(db),
                          u: User = Depends(require("vendor.edit"))):
        _RBAC.assert_object_visible(s, u, 'vendor', vid)
        v = s.scalars(select(VendorRecord).where(VendorRecord.vendor_id == vid)).first()
        if not v:
            raise HTTPException(404, "vendor not found")
        v.group_id = b.group_id
        audit(s, "v2.vendor_group_override", u.username, {"vendor_id": vid, "group_id": b.group_id})
        s.commit()
        return {"vendor_id": vid, "group_id": b.group_id}

    @app.post("/api/v2/contacts")
    def v2_add_contact(b: V2ContactIn, s: Session = Depends(db),
                       u: User = Depends(require("vendor.edit"))):
        c = RS.add_contact(s, owner_type=b.owner_type, owner_id=b.owner_id, name=b.name,
                           is_primary=b.is_primary, email=b.email,
                           phone_country_code=b.phone_country_code, phone_number=b.phone_number,
                           designation=b.designation, country=b.country,
                           mailing_address=b.mailing_address)
        audit(s, "v2.contact_added", u.username,
              {"owner": b.owner_id, "primary": b.is_primary, "contact_id": c.id})
        s.commit()
        return {"contact_id": c.id, "is_primary": c.is_primary}

    # ---- engagements (exhaustive) ----
    @app.post("/api/v2/engagements")
    def v2_create_engagement(b: V2EngagementIn, s: Session = Depends(db),
                             u: User = Depends(require("engagement.edit"))):
        e = RS.create_engagement(s, vendor_id=b.vendor_id, title=b.title,
                                 owner_user=b.owner_user or u.username,
                                 service_description=b.service_description,
                                 material_group_id=b.material_group_id,
                                 business_unit=b.business_unit,
                                 deployment_model=b.deployment_model,
                                 annual_value=b.annual_value,
                                 currency=b.currency)
        audit(s, "v2.engagement_created", u.username,
              {"engagement_id": e.engagement_id, "vendor_id": b.vendor_id})
        s.commit()
        from app.features.domain import watchlist_service as _WL
        wl = _WL.is_watchlisted(s, b.vendor_id)
        if wl:
            audit(s, "engagement.watchlist_signoff_required", u.username,
                  {"engagement_id": e.engagement_id, "vendor_id": b.vendor_id})
            s.commit()
        return {"engagement_id": e.engagement_id,
                "watchlist_signoff_required": wl,
                "notice": ("Supplier is on the watchlist — this engagement requires human "
                           "sign-off irrespective of risk level.") if wl else None}

    @app.get("/api/v2/engagements")
    def v2_list_engagements(vendor_id: Optional[str] = None, status: Optional[str] = None,
                            slim: bool = False,
                            s: Session = Depends(db), u: User = Depends(require("engagement.view"))):
        from datetime import date as _date
        _today = _date.today().isoformat()
        stmt = select(EngagementRecord)
        if vendor_id:
            stmt = stmt.where(EngagementRecord.vendor_id == vendor_id)
        _av = _RBAC.scoped_vendor_ids(s, u)
        if _av is not None:
            stmt = stmt.where(EngagementRecord.vendor_id.in_(_av or ["__none__"]))
        if status:
            stmt = stmt.where(EngagementRecord.status == status)
        rows = s.scalars(stmt).all()
        out = []
        for e in rows:
            if slim:
                # The list view in the SPA uses only these six fields.
                out.append({"engagement_id": e.engagement_id, "vendor_id": e.vendor_id,
                            "title": e.title, "status": e.status,
                            "residual_band": e.residual_band, "open_actions": e.open_actions})
                continue
            out.append({"engagement_id": e.engagement_id, "vendor_id": e.vendor_id,
                        "title": e.title, "status": e.status, "stage": e.stage,
                        "inherent_band": e.inherent_band, "residual_band": e.residual_band,
                        "open_actions": e.open_actions, "contract_id": e.contract_id,
                        "assessment_id": e.assessment_id, "material_group_id": e.material_group_id,
                        "last_assessment_date": e.last_assessment_date,
                        "next_assessment_due": e.next_assessment_due,
                        "reassessment_due": bool(e.next_assessment_due and e.next_assessment_due <= _today)})
        return out

    # ---- assessments ----
    @app.post("/api/v2/assessments")
    def v2_create_assessment(b: V2AssessmentIn, s: Session = Depends(db),
                             u: User = Depends(require("engagement.view"))):
        pool = [x.username for x in s.scalars(select(User)).all() if x.is_active]
        rec = RS.create_assessment(s, engagement_id=b.engagement_id, vendor_id=b.vendor_id,
                                   engagement_owner=u.username, session_id=b.session_id,
                                   inherent_band=b.inherent_band, residual_band=b.residual_band,
                                   assessor_pool=pool)
        audit(s, "v2.assessment_created", u.username,
              {"assessment_id": rec.assessment_id, "assessor": rec.assessor_user})
        s.commit()
        return {"assessment_id": rec.assessment_id, "status": rec.status,
                "assessor_user": rec.assessor_user, "spoc_user": rec.spoc_user}

    @app.get("/api/v2/assessments")
    def v2_list_assessments(s: Session = Depends(db), u: User = Depends(require("engagement.view"))):
        from sqlalchemy.orm import defer
        out = []
        # Defer the heavy structured_json payload — the list view never uses it.
        _astmt = select(AssessmentRecord).options(defer(AssessmentRecord.structured_json))
        _av = _RBAC.scoped_vendor_ids(s, u)
        if _av is not None:
            _astmt = _astmt.where(AssessmentRecord.vendor_id.in_(_av or ["__none__"]))
        for a in s.scalars(_astmt).all():
            if not _can_view_assessment(u, a):
                continue
            out.append({"assessment_id": a.assessment_id, "engagement_id": a.engagement_id,
                        "status": a.status, "inherent_band": a.inherent_band,
                        "outcome": a.outcome, "assessor_user": a.assessor_user,
                        "assessor_signed_off": a.assessor_signed_off, "locked": a.locked,
                        "spoc_user": a.spoc_user})
        return out

    @app.post("/api/v2/assessments/{aid}/signoff")
    def v2_signoff(aid: str, s: Session = Depends(db), u: User = Depends(require("engagement.review"))):
        _RBAC.assert_object_visible(s, u, 'assessment', aid)
        a = s.scalars(select(AssessmentRecord).where(AssessmentRecord.assessment_id == aid)).first()
        if not a:
            raise HTTPException(404, "assessment not found")
        if a.assessor_user and a.assessor_user != u.username and u.role.key != "admin":
            raise HTTPException(403, "only the assigned assessor may sign off")
        a.assessor_signed_off = True
        if a.engagement_id:
            RS.record_last_assessment(s, a.engagement_id)
        audit(s, "v2.assessment_signoff", u.username, {"assessment_id": aid})
        s.commit()
        return {"assessment_id": aid, "assessor_signed_off": True}

    @app.post("/api/v2/assessments/{aid}/approve")
    def v2_approve(aid: str, s: Session = Depends(db), u: User = Depends(require("engagement.review"))):
        _RBAC.assert_object_visible(s, u, 'assessment', aid)
        try:
            rec = RS.approve_assessment(s, aid)
        except ValueError as e:
            raise HTTPException(400, str(e))
        if rec and rec.engagement_id:
            RS.record_last_assessment(s, rec.engagement_id)
        audit(s, "v2.assessment_approved", u.username, {"assessment_id": aid, "locked": True})
        s.commit()
        return {"assessment_id": aid, "status": rec.status, "locked": rec.locked}

    @app.post("/api/v2/assessments/{aid}/recall")
    def v2_recall(aid: str, s: Session = Depends(db), u: User = Depends(require("engagement.view"))):
        _RBAC.assert_object_visible(s, u, 'assessment', aid)
        a = s.scalars(select(AssessmentRecord).where(AssessmentRecord.assessment_id == aid)).first()
        if not a:
            raise HTTPException(404, "assessment not found")
        if a.locked:
            raise HTTPException(400, "approved assessments are hard-locked and cannot be recalled")
        a.status = "Recalled"
        audit(s, "v2.assessment_recalled", u.username, {"assessment_id": aid})
        s.commit()
        return {"assessment_id": aid, "status": "Recalled"}

    @app.post("/api/v2/assessments/{aid}/reassign")
    def v2_reassign(aid: str, b: ReassignIn, s: Session = Depends(db),
                    u: User = Depends(require("engagement.review"))):
        _RBAC.assert_object_visible(s, u, 'assessment', aid)
        a = s.scalars(select(AssessmentRecord).where(AssessmentRecord.assessment_id == aid)).first()
        if not a:
            raise HTTPException(404, "assessment not found")
        if a.locked:
            raise HTTPException(400, "approved assessment is locked")
        a.assessor_user = b.assessor_user
        a.assessor_signed_off = False
        audit(s, "v2.assessor_reassigned", u.username, {"assessment_id": aid, "to": b.assessor_user})
        s.commit()
        return {"assessment_id": aid, "assessor_user": b.assessor_user}

    # ---- findings + remediation ----
    from app.features.domain.vocab import FINDING_STATUSES  # DB-05: single vocabulary

    @app.get("/api/v2/remediations")
    def v2_remediations_list(status: Optional[str] = None, s: Session = Depends(db),
                             u: User = Depends(require("finding.view"))):
        from app.features.domain.registry_models import RemediationRecord
        rows = s.scalars(select(RemediationRecord).order_by(RemediationRecord.id.desc())).all()
        return [_rmd_row(s, r) for r in rows if not status or r.status == status]

    @app.get("/api/v2/remediations/{rid}")
    def v2_remediation_get(rid: str, s: Session = Depends(db),
                           u: User = Depends(require("finding.view"))):
        from app.features.domain.registry_models import RemediationRecord
        r = s.scalars(select(RemediationRecord).where(RemediationRecord.remediation_id == rid)).first()
        if not r: raise HTTPException(404, "remediation plan not found")
        return _rmd_row(s, r)

    @app.put("/api/v2/remediations/{rid}")
    def v2_remediation_update(rid: str, body: dict = Body(...), s: Session = Depends(db),
                              u: User = Depends(require("finding.manage"))):
        from app.features.domain.registry_models import RemediationRecord
        from datetime import date as _date
        r = s.scalars(select(RemediationRecord).where(RemediationRecord.remediation_id == rid)).first()
        if not r: raise HTTPException(404, "remediation plan not found")
        for k in ("plan", "owner", "target_date", "status", "evidence", "verified_by"):
            if body.get(k) is not None:
                setattr(r, k, body[k])
        if body.get("progress_pct") is not None:
            r.progress_pct = max(0, min(100, int(body["progress_pct"])))
        if r.status in ("Complete", "Verified") and not r.completed_date:
            r.completed_date = _date.today().isoformat()
            r.progress_pct = 100
        audit(s, "v2.remediation_update", u.username, {"remediation_id": rid, "status": r.status})
        s.commit()
        return _rmd_row(s, r)

    @app.post("/api/v2/engagements/{eid}/assign")
    def v2_assign_engagement(eid: str, b: AssignAssessorIn, s: Session = Depends(db),
                             u: User = Depends(require("engagement.assign"))):
        _RBAC.assert_object_visible(s, u, 'engagement', eid)
        eng = s.scalars(select(EngagementRecord).where(
            EngagementRecord.engagement_id == eid)).first()
        if not eng: raise HTTPException(404, "engagement not found")
        target = s.scalars(select(User).where(User.username == b.assessor_user)).first()
        if not target: raise HTTPException(404, "assessor user not found")
        eng.assigned_assessor = b.assessor_user
        # propagate to the engagement's assessments + open findings
        for a in s.scalars(select(AssessmentRecord).where(
                AssessmentRecord.engagement_id == eid)).all():
            a.assessor_user = b.assessor_user
        for f in s.scalars(select(FindingRecord).where(
                FindingRecord.engagement_id == eid)).all():
            if not f.assessor:
                f.assessor = b.assessor_user
        s.flush()
        audit(s, "v2.engagement_assigned", u.username,
              {"engagement_id": eid, "assessor": b.assessor_user})
        notify(s, f"Engagement {eid} assigned to {b.assessor_user}", "vrm")
        s.commit()
        return {"engagement_id": eid, "assigned_assessor": b.assessor_user}

    @app.get("/api/v2/assessors")
    def v2_list_assessors(s: Session = Depends(db),
                          u: User = Depends(require("engagement.assign"))):
        from app.features.domain.models_db import Role
        rid = s.scalars(select(Role).where(Role.key == "vrm")).first()
        users = s.scalars(select(User).where(User.role_id == (rid.id if rid else -1))).all()
        return [{"username": x.username} for x in users]

    # ===== Continuous-monitoring connectors: RapidRatings (FDD) + Interos (Reputation) =====
    # API trigger + inbound webhook + MCP-style discovery. Demonstrator data with a clean
    # seam: set the provider credentials/MCP endpoint to switch to live calls.
    @app.post("/api/v2/webhooks/rapidratings")
    def v2_webhook_rapidratings(b: ConnectorWebhookIn, s: Session = Depends(db),
                                u: User = Depends(require("admin.webhooks"))):
        payload = b.payload or {"provider": "RapidRatings", "event": "rating_change"}
        doc = _file_monitoring_report(s, b.vendor_id, "rapidratings", "webhook", payload, u.username)
        audit(s, "v2.webhook_rapidratings", u.username, {"vendor_id": b.vendor_id})
        s.commit()
        return {"received": True, "filed_report": doc}

    @app.post("/api/v2/webhooks/interos")
    def v2_webhook_interos(b: ConnectorWebhookIn, s: Session = Depends(db),
                           u: User = Depends(require("admin.webhooks"))):
        payload = b.payload or {"provider": "Interos", "event": "risk_event"}
        doc = _file_monitoring_report(s, b.vendor_id, "interos", "webhook", payload, u.username)
        audit(s, "v2.webhook_interos", u.username, {"vendor_id": b.vendor_id})
        s.commit()
        return {"received": True, "filed_report": doc}

    # ===== Global Regulations — regulatory reference, AI live-updates & gap assessment =====
    import json as _json2, os as _os2
    _REG_CACHE = {}

    def _regcat():
        if not _REG_CACHE:
            # regdata.json lives in app/features/. This file moved to app/routers/
            # in the router split, so resolve from the app package root (with the
            # old sibling path kept as a fallback for safety).
            _base = _os2.path.dirname(_os2.path.dirname(_os2.path.abspath(__file__)))
            for path in (_os2.path.join(_base, "features", "regdata.json"),
                         _os2.path.join(_os2.path.dirname(_os2.path.abspath(__file__)),
                                        "features", "regdata.json")):
                if _os2.path.exists(path):
                    with open(path) as f:
                        _REG_CACHE.update(_json2.load(f))
                    break
            else:
                raise HTTPException(503, "regulations catalog (regdata.json) not found")
        return _REG_CACHE

    def _reg_short(t):
        return (str(t).split("\u2014")[0].split("(")[0].strip())[:60]

    def _extract_json_array(text):
        if not text:
            return []
        t = text.replace("```json", "").replace("```", "").strip()
        st, e = t.find("["), t.rfind("]")
        if st == -1 or e == -1 or e < st:
            return []
        try:
            arr = json.loads(t[st:e + 1])
            return [x for x in arr if isinstance(x, dict)] if isinstance(arr, list) else []
        except Exception:
            return []

    def _reg_norm_status(sv):
        t = str(sv or "").lower()
        if t.startswith("add") or t in ("full", "green", "yes"):
            return "addressed"
        if t.startswith("part") or t in ("amber", "some"):
            return "partial"
        return "gap"

    @app.get("/api/v2/regulations")
    def v2_regulations(u: User = Depends(require("reg.view"))):
        d = _regcat()
        return {"attrs": d["ATTRS"], "order": d["ORDER"], "extra_order": d["EXTRA"],
                "industries": d["INDUSTRIES"], "catalog": d["CATALOG"]}

    @app.post("/api/v2/regulations/export")
    def v2_reg_export(b: RegExportIn, s: Session = Depends(db),
                      u: User = Depends(require("reg.view"))):
        from openpyxl import Workbook
        from fastapi import Response
        import io
        d = _regcat(); cat = d["CATALOG"]; ATTRS = d["ATTRS"]
        wb = Workbook(); wb.remove(wb.active)
        for code in b.codes:
            c = cat.get(code)
            if not c:
                continue
            ws = wb.create_sheet(code[:31])
            ws.append(["Attribute"] + [_reg_short(inst[0]) for inst in c["I"]])
            for ai, attr in enumerate(ATTRS):
                ws.append([attr] + [(inst[ai] if ai < len(inst) else "") for inst in c["I"]])
        if b.updates:
            ws = wb.create_sheet("Live Updates")
            ws.append(["Jurisdiction", "Instrument idx", "Attribute idx", "Title", "Update", "Date", "Source"])
            for up in b.updates:
                ws.append([up.get("full") or up.get("code"), up.get("instrument"),
                           up.get("attr"), up.get("title"), up.get("update"),
                           up.get("date"), up.get("source")])
        if not wb.sheetnames:
            wb.create_sheet("Empty")
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        audit(s, "v2.reg_export", u.username, {"codes": b.codes}); s.commit()
        return Response(content=buf.read(),
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        headers={"Content-Disposition": 'attachment; filename="Global_Regulations.xlsx"'})

    @app.post("/api/v2/regulations/refresh")
    def v2_reg_refresh(b: RegRefreshIn, s: Session = Depends(db),
                       u: User = Depends(require("reg.assess"))):
        from app.agents import llm_config
        if not llm_config.status().get("live_ready"):
            return {"holding": True, "message": AI_HOLDING, "updates": []}
        d = _regcat(); cat = d["CATALOG"]; ATTRS = d["ATTRS"]; updates = []
        attr_idx = "|".join(f"{i}:{a}" for i, a in enumerate(ATTRS))
        for code in (b.codes or [])[:6]:
            c = cat.get(code)
            if not c:
                continue
            inst_idx = "|".join(f"{i}:{_reg_short(inst[0])}" for i, inst in enumerate(c["I"])) or "(none yet)"
            prompt = "\n".join([
                f"FS analyst. {c['full']} ({c['reg']}).",
                f"Instruments: {inst_idx}", f"Attrs: {attr_idx}",
                "Web-search 2026-onward TPRM/outsourcing/ICT-resilience news: rules, consultations,",
                "regulator letters, enforcement, fines, CTP designations. Primary sources only.",
                "Map each to instrument+attr index. New instrument: instrument:-1.",
                "JSON array only, no prose/fences.",
                "Per item: {instrument:int,attr:int,title:str,update:str<=180c,date:str,source:str}.",
                "Max 6 newest-first. Nothing new: []."])
            try:
                out = llm_config.complete(
                    PROMPTS.resolve(s, "regulatory_websearch"),
                    prompt, domain="regulatory", web_search=True, max_tokens=900)
                for it in _extract_json_array(out):
                    it["code"] = code; it["full"] = c["full"]; updates.append(it)
            except Exception:
                continue
        audit(s, "v2.reg_refresh", u.username, {"codes": b.codes, "found": len(updates)})
        s.commit()
        return {"holding": False, "updates": updates}

    @app.post("/api/v2/regulations/assess")
    def v2_reg_assess(b: RegAssessIn, s: Session = Depends(db),
                      u: User = Depends(require("reg.assess"))):
        from app.agents import llm_config
        if not llm_config.status().get("live_ready"):
            return {"holding": True, "message": AI_HOLDING, "results": {}, "coverage": {}}
        d = _regcat(); cat = d["CATALOG"]; results = {}
        tally = {"addressed": 0, "partial": 0, "gap": 0}
        doc = (b.doc_text or "")[:8000]
        for code in (b.codes or []):
            c = cat.get(code)
            if not c or not c["I"]:
                continue
            regs = "\n".join(
                f"{i}|{_reg_short(inst[0])}|clauses:{(inst[7] or '—')[:110]}|clock:{inst[8] or '—'}"
                for i, inst in enumerate(c["I"]))
            prompt = "\n".join([
                f"Compliance assessor, {c['full']}.",
                "Rate each regulation vs docs: addressed=substantially covers, partial=some gaps, gap=not covered.",
                "Conservative: absent obligation = gap.", f"REGS:\n{regs}",
                f'DOCS:\n"""\n{doc}\n"""', "JSON array only, no prose/fences.",
                "Per item: {instrument:int,status:str,rationale:str<=160c,gap:str<=120c}."])
            try:
                out = llm_config.complete(
                    PROMPTS.resolve(s, "compliance_assessor"),
                    prompt, domain="regulatory", max_tokens=1200)
                items = []
                for it in _extract_json_array(out):
                    st = _reg_norm_status(it.get("status")); it["status"] = st
                    ii = it.get("instrument")
                    if isinstance(ii, int) and 0 <= ii < len(c["I"]):
                        it["title"] = _reg_short(c["I"][ii][0])
                    tally[st] = tally.get(st, 0) + 1; items.append(it)
                results[code] = {"full": c["full"], "items": items}
            except Exception:
                continue
        audit(s, "v2.reg_assess", u.username, {"codes": b.codes}); s.commit()
        return {"holding": False, "results": results, "coverage": tally}

    @app.post("/api/v2/regulations/relevance")
    def v2_reg_relevance(b: RegRelevanceIn, s: Session = Depends(db),
                         u: User = Depends(require("reg.assess"))):
        from app.agents import llm_config
        if not llm_config.status().get("live_ready"):
            return {"holding": True, "message": AI_HOLDING, "relevance": {}}
        d = _regcat(); cat = d["CATALOG"]; rel = {}
        ind = b.industry or {}
        ilabel = ind.get("other") if ind.get("id") == "other" else \
            next((x["label"] for x in d["INDUSTRIES"] if x["id"] == ind.get("id")), "All")
        for code in (b.codes or []):
            c = cat.get(code)
            if not c or not c["I"]:
                continue
            lst = "\n".join(f"{k}:{_reg_short(inst[0])}|{(inst[5] or '')[:80]}"
                            for k, inst in enumerate(c["I"]))
            prompt = "\n".join([
                f'Industry: "{ilabel}". Jurisdiction: {c["full"]}.',
                "Mark applies:false ONLY when instrument clearly does NOT bind this industry.",
                "Banks: all. Trading houses: market-conduct/data/resilience, not deposit-taking rules. FS: most.",
                f"INSTRUMENTS:\n{lst}", "JSON array only, no prose/fences.",
                "Per item: {instrument:int,applies:bool,note:str<=100c}."])
            try:
                out = llm_config.complete(PROMPTS.resolve(s, "regulatory_applicability"),
                                          prompt, domain="regulatory", max_tokens=900)
                rel[code] = _extract_json_array(out)
            except Exception:
                continue
        audit(s, "v2.reg_relevance", u.username, {"codes": b.codes}); s.commit()
        return {"holding": False, "relevance": rel}

    @app.post("/api/v2/regulations/assess/export")
    def v2_reg_assess_export(b: RegReportExportIn, s: Session = Depends(db),
                             u: User = Depends(require("reg.view"))):
        from openpyxl import Workbook
        from fastapi import Response
        import io
        rep = b.report or {}; lab = {"addressed": "Addressed", "partial": "Partial", "gap": "Gap"}
        wb = Workbook(); ws = wb.active; ws.title = "Assessment"
        ws.append([f"Global Regulations — Documentation Assessment"])
        ws.append(["Jurisdiction", "Instrument", "Status", "Rationale", "Gap"])
        for code, r in (rep.get("results") or {}).items():
            for it in (r.get("items") or []):
                ws.append([r.get("full", code), it.get("title", ""),
                           lab.get(it.get("status"), it.get("status", "")),
                           it.get("rationale", ""), it.get("gap", "")])
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        audit(s, "v2.reg_assess_export", u.username, {}); s.commit()
        return Response(content=buf.read(),
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        headers={"Content-Disposition": 'attachment; filename="Regulatory_Assessment.xlsx"'})

    # ===== Phase 1 — Data Integrity & Completion (the overnight steward) =====
    _ALLOWED_FIX_FIELDS = {"lei", "registration_number", "incorporation_country",
                           "hq_country", "ultimate_parent", "legal_form", "website",
                           "listing_status", "status", "tier", "trading_name", "duns"}

    @app.get("/api/v2/evidence/{vendor_id}")
    def v2_evidence(vendor_id: str, s: Session = Depends(db),
                    u: User = Depends(require("vendor.view"))):
        _RBAC.assert_object_visible(s, u, "vendor", vendor_id)
        from app.features.domain.registry_models import (VendorRecord, EngagementRecord,
                                              AssessmentRecord, FindingRecord, IncidentRecord)
        v = s.scalars(select(VendorRecord).where(VendorRecord.vendor_id == vendor_id)).first()
        if not v:
            raise HTTPException(404, "vendor not found")
        engs = s.scalars(select(EngagementRecord).where(EngagementRecord.vendor_id == vendor_id)).all()
        asms = s.scalars(select(AssessmentRecord).where(AssessmentRecord.vendor_id == vendor_id)).all()
        finds = s.scalars(select(FindingRecord).where(FindingRecord.vendor_id == vendor_id)).all()
        incs = s.scalars(select(IncidentRecord).where(IncidentRecord.vendor_id == vendor_id)).all()
        docs = []
        try:
            from app.features.lifecycle.documents import StoredDocument as _SD
            docs = s.scalars(select(_SD).where(_SD.vendor_id == vendor_id)).all()
        except Exception:
            docs = []
        ids = {vendor_id} | {e.engagement_id for e in engs} | {a.assessment_id for a in asms} \
            | {f.finding_id for f in finds} | {i.incident_id for i in incs}
        # Audit trail for this subject (DB-03): served by the indexed subject columns
        # instead of loading the entire chain and string-matching the JSON payload.
        # Legacy rows written before the subject columns existed are still found via
        # a bounded payload match, so history stays visible during backfill.
        rel = s.scalars(select(AuditLog)
                        .where(AuditLog.vendor_id == vendor_id)
                        .order_by(AuditLog.seq.desc()).limit(500)).all()
        if len(rel) < 500:
            seen = {r.seq for r in rel}
            legacy = s.scalars(select(AuditLog)
                               .where(AuditLog.vendor_id.is_(None))
                               .order_by(AuditLog.seq.desc()).limit(2000)).all()
            for r in legacy:
                if r.seq not in seen and r.detail and any(i in r.detail for i in ids):
                    rel.append(r)
        rel.sort(key=lambda r: r.seq)
        # Chain integrity is a whole-chain property and is reported by
        # /api/v1/audit/verify; re-walking it inside a vendor page does not scale.
        intact = None
        approvals = [{"assessment_id": a.assessment_id, "outcome": a.outcome,
                      "assessor": a.assessor_user, "signed_off": a.assessor_signed_off,
                      "locked": a.locked, "status": a.status} for a in asms]
        signed = sum(1 for a in asms if a.assessor_signed_off)
        return {
            "vendor": {"vendor_id": v.vendor_id, "legal_name": v.legal_name, "tier": v.tier,
                       "is_critical": v.is_critical, "hq_country": v.hq_country,
                       "lei": v.lei, "status": v.status},
            "engagements": [{"engagement_id": e.engagement_id, "title": e.title,
                             "inherent_band": e.inherent_band, "residual_band": e.residual_band,
                             "business_unit": e.business_unit,
                             "annual_value": round(float(e.annual_value or 0))} for e in engs],
            "assessments": [{"assessment_id": a.assessment_id, "status": a.status,
                             "outcome": a.outcome, "inherent_band": a.inherent_band,
                             "residual_band": a.residual_band, "assessor": a.assessor_user,
                             "signed_off": a.assessor_signed_off, "locked": a.locked,
                             "created_at": a.created_at.isoformat() if a.created_at else None} for a in asms],
            "findings": [{"finding_id": f.finding_id, "title": f.title, "severity": f.severity,
                          "status": f.status, "source": f.source,
                          "risk_accepted": getattr(f, "risk_accepted", False),
                          "acceptance_expiry": getattr(f, "acceptance_expiry", None)} for f in finds],
            "documents": [{"doc_id": d.doc_id, "filename": d.filename, "purpose": d.purpose,
                           "uploaded_by": d.uploaded_by,
                           "created_at": d.created_at.isoformat() if getattr(d, "created_at", None) else None} for d in docs],
            "incidents": [{"incident_id": i.incident_id, "incident_type": i.incident_type,
                           "severity": i.severity, "status": i.status,
                           "notification_compliant": i.notification_compliant} for i in incs],
            "approvals": approvals,
            "audit_trail": [{"seq": r.seq, "action": r.action, "actor": r.actor,
                             "created_at": r.created_at.isoformat() if r.created_at else None,
                             "entry_hash": (r.entry_hash or "")[:16]} for r in rel[-40:]],
            "summary": {"engagements": len(engs), "assessments": len(asms),
                        "signed_off": signed, "findings": len(finds),
                        "documents": len(docs), "incidents": len(incs),
                        "audit_entries": len(rel)},
            "chain_intact": intact,
            "chain_note": "Chain integrity is a whole-chain property — see /api/v1/audit/verify.",
        }

    @app.get("/api/v2/board-pack")
    def v2_board_pack(u: User = Depends(require("vendor.view")), s: Session = Depends(db)):
        from app.features.intelligence import exposure as EXP
        from app.features.intelligence import graph as GRAPH
        from app.features.assessment import criticality as CRIT
        from app.features.intelligence import geopolitical as GEO
        from app.features.lifecycle import integrity as INTEG
        from app.features.domain.registry_models import (VendorRecord, EngagementRecord,
                                              FindingRecord, IncidentRecord)
        from app.agents import llm_config
        vendors = s.scalars(select(VendorRecord)).all()
        engs = s.scalars(select(EngagementRecord)).all()
        finds = s.scalars(select(FindingRecord)).all()
        incs = s.scalars(select(IncidentRecord).order_by(IncidentRecord.id.desc())).all()
        open_find = [f for f in finds if f.status != "Closed"]
        hi_find = [f for f in open_find if (f.severity or "") in ("High", "Critical")]
        exec_stats = {
            "vendors": len(vendors), "critical_vendors": sum(1 for v in vendors if v.is_critical),
            "engagements": len(engs), "open_findings": len(open_find),
            "high_critical_findings": len(hi_find),
            "annual_value": round(sum(float(e.annual_value or 0) for e in engs)),
        }
        graph = GRAPH.build_graph(s)
        bu = EXP.bu_exposure(s)
        crit = CRIT.model(s)
        geo = GEO.exposure(s)
        dq = INTEG.run_sweep(s)["health"]
        inc_open = [i for i in incs if i.status != "Closed"]
        inc_sev = {}
        for i in incs:
            inc_sev[i.severity] = inc_sev.get(i.severity, 0) + 1
        inc_breach = sum(1 for i in incs if i.notification_compliant == "red")
        disagreements = [c for c in crit["subthreshold_high_risk"] if c.get("model_critical")]
        payload = {
            "generated_at": __import__("datetime").datetime.now(__import__("datetime").timezone.utc).isoformat(timespec="seconds"),
            "exec": exec_stats,
            "concentration": {"shared_sub_providers": graph["stats"]["shared_fourth_parties"],
                              "spof": graph["stats"]["spof_count"],
                              "max_fanout_pct": graph["stats"]["max_fanout_pct"],
                              "top": graph["shared_fourth_parties"][:5]},
            "bu_exposure": bu["business_units"][:6],
            "criticality": {"critical": crit["stats"]["critical"],
                            "disagreements": len(disagreements),
                            "top_disagreements": disagreements[:5]},
            "geopolitical": {"exposed": geo["stats"]["exposed_vendors"],
                             "component_shortage": geo["stats"]["component_shortage_vendors"],
                             "high_risk_jurisdictions": geo["stats"]["high_risk_jurisdictions"],
                             "top": geo["flagged_vendors"][:6]},
            "incidents": {"total": len(incs), "open": len(inc_open),
                          "by_severity": inc_sev, "notification_breaches": inc_breach,
                          "recent": [{"incident_id": i.incident_id, "vendor_name": i.vendor_name,
                                      "severity": i.severity, "status": i.status,
                                      "incident_type": i.incident_type,
                                      "notification_compliant": i.notification_compliant}
                                     for i in incs[:6]]},
            "data_health": {"overall": dq["overall"], "completeness": dq["completeness"],
                            "issues": dq["issue_count"]},
        }
        # executive summary — deterministic, AI-enhanced if live
        summary = (f"The third-party estate comprises {exec_stats['vendors']} vendors "
                   f"({exec_stats['critical_vendors']} critical) across {exec_stats['engagements']} engagements, "
                   f"£{exec_stats['annual_value']:,} annual value. {exec_stats['high_critical_findings']} high/critical "
                   f"findings are open. Concentration: {graph['stats']['spof_count']} single points of failure, the "
                   f"largest serving {graph['stats']['max_fanout_pct']}% of the estate. {len(incs)} incidents "
                   f"({len(inc_open)} open, {inc_breach} with a notification-SLA breach). {geo['stats']['exposed_vendors']} "
                   f"vendors carry geopolitical/export-control exposure. Data-health score {dq['overall']}/100.")
        payload["summary_ai"] = False
        if llm_config.status().get("live_ready"):
            try:
                out = llm_config.complete(
                    PROMPTS.resolve(s, "board_pack_summary"),
                    summary, domain="risk", max_tokens=500)
                if out and out.strip():
                    summary = out.strip(); payload["summary_ai"] = True
            except Exception as _e:
                _obs_swallow('bro_app.py', _e)
        payload["summary"] = summary
        return payload

    @app.post("/api/v2/remediations")
    def v2_create_remediation(b: V2RemediationIn, s: Session = Depends(db),
                              u: User = Depends(require("finding.manage"))):
        r = RS.create_remediation(s, finding_id=b.finding_id, plan=b.plan,
                                  owner=b.owner, target_date=b.target_date)
        audit(s, "v2.remediation_created", u.username,
              {"remediation_id": r.remediation_id, "finding_id": b.finding_id})
        s.commit()
        return {"remediation_id": r.remediation_id}

    # ---- fourth parties ----
    @app.post("/api/v2/fourth-parties")
    def v2_create_fourth_party(b: V2FourthPartyIn, s: Session = Depends(db),
                               u: User = Depends(require("lifecycle.fourthparty"))):
        fp = RS.create_fourth_party(s, legal_name=b.legal_name, vendor_ids=b.vendor_ids or [],
                                    also_vendor_id=b.vendor_id, service_provided=b.service_provided,
                                    hq_country=b.hq_country)
        audit(s, "v2.fourth_party_created", u.username,
              {"fourth_party_id": fp.fourth_party_id, "concentration": fp.concentration_flag})
        s.commit()
        return {"fourth_party_id": fp.fourth_party_id, "concentration_flag": fp.concentration_flag}

    @app.get("/api/v2/fourth-parties")
    def v2_list_fourth_parties(s: Session = Depends(db), u: User = Depends(require("lifecycle.fourthparty"))):
        out = []
        for fp in s.scalars(select(FourthPartyRecord)).all():
            vlinks = [fv.vendor_id for fv in s.scalars(select(FourthPartyVendor).where(
                FourthPartyVendor.fourth_party_id == fp.fourth_party_id)).all()]
            out.append({"fourth_party_id": fp.fourth_party_id, "legal_name": fp.legal_name,
                        "concentration_flag": fp.concentration_flag, "vendor_id": fp.vendor_id,
                        "supports_vendors": vlinks})
        return out

    @app.get("/api/v2/fourth-parties/{fpid}/vendors")
    def v2_fourth_party_vendors(fpid: str, s: Session = Depends(db),
                                u: User = Depends(require("lifecycle.fourthparty"))):
        from app.features.domain.registry_models import VendorRecord
        fp = s.scalars(select(FourthPartyRecord).where(FourthPartyRecord.fourth_party_id == fpid)).first()
        if not fp:
            raise HTTPException(404, "fourth party not found")
        vids = [fv.vendor_id for fv in s.scalars(select(FourthPartyVendor).where(
            FourthPartyVendor.fourth_party_id == fpid)).all()]
        vmap = {v.vendor_id: v for v in s.scalars(select(VendorRecord).where(
            VendorRecord.vendor_id.in_(vids))).all()} if vids else {}
        vendors = [{"vendor_id": vid, "legal_name": (vmap[vid].legal_name if vid in vmap else vid),
                    "tier": (vmap[vid].tier if vid in vmap else None),
                    "is_critical": (bool(vmap[vid].is_critical) if vid in vmap else False)}
                   for vid in vids]
        return {"fourth_party_id": fpid, "legal_name": fp.legal_name,
                "concentration_flag": fp.concentration_flag,
                "supports_count": len(vendors), "vendors": vendors}

    # ---- artefacts + revalidation + issues ----
    @app.post("/api/v2/artefacts")
    def v2_create_artefact(b: V2ArtefactIn, s: Session = Depends(db),
                           u: User = Depends(require("lifecycle.documents"))):
        art = RS.create_artefact(s, vendor_id=b.vendor_id, name=b.name,
                                 artefact_type=b.artefact_type or "certificate",
                                 expiry_date=b.expiry_date, received_via=b.received_via or "upload",
                                 supersedes=b.supersedes, issue_date=b.issue_date,
                                 engagement_id=b.engagement_id)
        audit(s, "v2.artefact_created", u.username,
              {"artefact_id": art.artefact_id, "status": art.status})
        s.commit()
        return {"artefact_id": art.artefact_id, "status": art.status,
                "supersedes": art.supersedes}

    @app.get("/api/v2/artefacts")
    def v2_list_artefacts(vendor_id: Optional[str] = None, s: Session = Depends(db),
                          u: User = Depends(require("lifecycle.documents"))):
        rows = s.scalars(select(ArtefactRecord)).all()
        return [{"artefact_id": a.artefact_id, "vendor_id": a.vendor_id, "name": a.name,
                 "type": a.artefact_type, "expiry_date": a.expiry_date, "status": a.status,
                 "is_current": a.is_current, "received_via": a.received_via,
                 "doc_link": a.object_uri}
                for a in rows if not vendor_id or a.vendor_id == vendor_id]

    # ============================================================
    # CR-4/5/12 — DOCUMENT STORE + AI EXTRACTION
    # ============================================================
    @app.post("/api/v2/certificates/ingest")
    def v2_cert_ingest(b: CertIngestIn, s: Session = Depends(db),
                       u: User = Depends(require("lifecycle.documents"))):
        """CR-5: multi-document certificate ingest. Each document is stored, read by the
        extractor, and a certificate record (ArtefactRecord) is created with the document
        linked for viewing, tagged to the vendor (and engagement where given)."""
        from app.features.lifecycle import documents as DOC
        v = s.scalars(select(VendorRecord).where(VendorRecord.vendor_id == b.vendor_id)).first()
        if not v:
            raise HTTPException(404, "vendor not found")
        created = []
        for f in b.files:
            try:
                doc = DOC.store_document(s, filename=f.filename, content_type=f.content_type or "",
                                         data_b64=f.data_b64, vendor_id=b.vendor_id,
                                         engagement_id=b.engagement_id, uploaded_by=u.username,
                                         purpose="certificate")
            except ValueError as e:
                raise HTTPException(422, str(e))
            ext = DOC.extract_certificate(s, doc)
            art = RS.create_artefact(
                s, vendor_id=b.vendor_id, name=ext["name"],
                artefact_type=ext["artefact_type"], expiry_date=ext.get("expiry_date"),
                received_via="upload", issue_date=ext.get("issue_date"),
                engagement_id=b.engagement_id, object_uri=f"/api/v2/documents/{doc.doc_id}")
            created.append({"artefact_id": art.artefact_id, "name": art.name,
                            "type": art.artefact_type, "expiry_date": art.expiry_date,
                            "status": art.status, "doc_id": doc.doc_id,
                            "doc_link": f"/api/v2/documents/{doc.doc_id}",
                            "gaps": ext.get("gaps", [])})
        audit(s, "v2.cert_ingest", u.username,
              {"vendor_id": b.vendor_id, "count": len(created)})
        s.commit()
        return {"certificates": created}

    @app.post("/api/v2/artefacts/revalidate")
    def v2_revalidate(s: Session = Depends(db), u: User = Depends(require("lifecycle.evidence"))):
        result = RS.revalidation_run(s)
        # send 7-day notices via the email path (SMTP or simulation)
        for n in result["notify_7day"]:
            v = s.scalars(select(VendorRecord).where(VendorRecord.vendor_id == n["vendor_id"])).first()
            contact = s.scalars(select(ContactRecord).where(
                ContactRecord.owner_type == "vendor", ContactRecord.owner_id == n["vendor_id"],
                ContactRecord.is_primary == True)).first()  # noqa: E712
            to = contact.email if contact and contact.email else "vendor@example.com"
            s.add(EmailOutbox(to_addr=to, subject=f"Certificate expiring: {n['name']}",
                              body=f"{n['name']} expires {n['expiry']}. Please provide an updated copy.",
                              sent=False))
        audit(s, "v2.revalidation_run", u.username,
              {"checked": result["checked"], "notify": len(result["notify_7day"]),
               "new_issues": len(result["new_issues"])})
        s.commit()
        return result

    @app.get("/api/v2/issues")
    def v2_list_issues(status: Optional[str] = None, s: Session = Depends(db),
                       u: User = Depends(require("finding.view"))):
        rows = s.scalars(select(IssueRecord).order_by(IssueRecord.id.desc())).all()
        from app.features.domain.registry_models import ArtefactRecord
        art_eng = {a.artefact_id: a.engagement_id for a in s.scalars(select(ArtefactRecord)).all()}
        return [{"issue_id": i.issue_id, "vendor_id": i.vendor_id, "vendor_name": i.vendor_name,
                 "engagement_id": art_eng.get(i.artefact_id),
                 "artefact_id": i.artefact_id, "kind": i.kind, "detail": i.detail,
                 "status": i.status, "closed_reason": i.closed_reason}
                for i in rows if not status or i.status == status]

    # ---- financial DD (deterministic engine) ----
    @app.get("/api/v2/config")
    def v2_config_list(s: Session = Depends(db), u: User = Depends(require("admin.config"))):
        from app.features.domain import config_store as CFG
        cats = {}
        for r in CFG.list_config(s):
            if r.value_type == "json" or r.category == "_internal" or r.key.startswith("nav."):
                continue
            eff = CFG.get_config(s, r.key)
            dflt = CFG.default_for(r.key)
            cats.setdefault(r.category, []).append({
                "key": r.key, "label": r.label, "description": r.description,
                "type": r.value_type, "value": eff, "default": dflt,
                "is_default": eff == dflt, "updated_by": r.updated_by,
                "updated_at": r.updated_at.isoformat() if r.updated_at else None})
        return {"categories": [{"category": c, "items": items} for c, items in cats.items()]}

    @app.put("/api/v2/config/{key}")
    def v2_config_set(key: str, b: ConfigSetIn, s: Session = Depends(db),
                      u: User = Depends(require("admin.config"))):
        from app.features.domain import config_store as CFG
        try:
            row = CFG.set_config(s, key, b.value, updated_by=u.username)
        except (ValueError, TypeError):
            raise HTTPException(status_code=400, detail="Invalid value for this setting type")
        if not row:
            raise HTTPException(status_code=404, detail="Unknown setting")
        audit(s, "v2.config_set", u.username, {"key": key, "value": CFG.get_config(s, key)})
        s.commit()
        return {"key": key, "value": CFG.get_config(s, key), "updated_by": row.updated_by}

    @app.post("/api/v2/config/{key}/reset")
    def v2_config_reset(key: str, s: Session = Depends(db),
                        u: User = Depends(require("admin.config"))):
        from app.features.domain import config_store as CFG
        row = CFG.reset_config(s, key, updated_by=u.username)
        if not row:
            raise HTTPException(status_code=404, detail="Unknown setting")
        audit(s, "v2.config_reset", u.username, {"key": key})
        s.commit()
        return {"key": key, "value": CFG.get_config(s, key), "is_default": True}

    # ---- navigation menu order (read: any user; write: admin) ----
    @app.get("/api/v2/nav-order")
    def v2_nav_order_get(s: Session = Depends(db), u: User = Depends(actor)):
        from app.features.domain import config_store as CFG
        return {"order": CFG.get_json(s, "nav.order", None)}

    @app.put("/api/v2/nav-order")
    def v2_nav_order_set(b: NavOrderIn, s: Session = Depends(db),
                         u: User = Depends(require("admin.config"))):
        from app.features.domain import config_store as CFG
        if b.order is None:
            CFG.delete_key(s, "nav.order")
            audit(s, "v2.nav_order_reset", u.username, {})
            s.commit()
            return {"order": None}
        CFG.upsert_json(s, "nav.order", b.order, updated_by=u.username)
        audit(s, "v2.nav_order_set", u.username,
              {"groups": len((b.order or {}).get("groups", []))})
        s.commit()
        return {"order": CFG.get_json(s, "nav.order", None)}

    # ---- contract management (Matt) ----
    @app.post("/api/v2/assessments/from-session")
    def v2_capture_session(b: CaptureIn, s: Session = Depends(db),
                           u: User = Depends(require("engagement.view"))):
        from app.features.assessment import assessment_capture as CAP
        pool = [x.username for x in s.scalars(select(User)).all() if x.is_active]
        try:
            rec = CAP.capture_session(s, session_id=b.session_id,
                                      engagement_id=b.engagement_id,
                                      vendor_id=b.vendor_id,
                                      engagement_owner=u.username, assessor_pool=pool)
        except ValueError as e:
            raise HTTPException(400, str(e))
        # link any documents uploaded during the conversation to this engagement/vendor
        # so they are retrievable against the assessment for audit
        try:
            from app.features.lifecycle.documents import StoredDocument as _SD
            for _d in s.scalars(select(_SD).where(
                    _SD.purpose == f"broassess:{b.session_id}")).all():
                _d.engagement_id = b.engagement_id
                if b.vendor_id:
                    _d.vendor_id = b.vendor_id
        except Exception:
            pass
        audit(s, "v2.assessment_captured", u.username,
              {"assessment_id": rec.assessment_id, "session_id": b.session_id,
               "status": rec.status})
        s.commit()
        return {"assessment_id": rec.assessment_id, "status": rec.status,
                "inherent_band": rec.inherent_band, "assessor_user": rec.assessor_user}

    @app.get("/api/v2/assessments/{aid}/structured")
    def v2_assessment_structured(aid: str, s: Session = Depends(db),
                                 u: User = Depends(require("engagement.view"))):
        _RBAC.assert_object_visible(s, u, 'assessment', aid)
        a = s.scalars(select(AssessmentRecord).where(
            AssessmentRecord.assessment_id == aid)).first()
        if not a:
            raise HTTPException(404, "assessment not found")
        if not _can_view_assessment(u, a):
            raise HTTPException(403, "you do not have access to this assessment record")
        return {"assessment_id": a.assessment_id, "engagement_id": a.engagement_id,
                "status": a.status, "locked": a.locked,
                "structured": json.loads(a.structured_json or "{}")}

    @app.get("/api/v2/assessments/{aid}/review")
    def v2_assessment_review(aid: str, s: Session = Depends(db),
                             u: User = Depends(require("engagement.view"))):
        _RBAC.assert_object_visible(s, u, 'assessment', aid)
        """CR-2: full reviewable detail — scope, inherent risks, controls assessed,
        documents and the final residual recommendation — for a reviewer to scrutinise
        before approving. CR-3 access rule enforced."""
        from app.features.domain import master_service as MS
        a = s.scalars(select(AssessmentRecord).where(
            AssessmentRecord.assessment_id == aid)).first()
        if not a:
            raise HTTPException(404, "assessment not found")
        if not _can_view_assessment(u, a):
            raise HTTPException(403, "you do not have access to this assessment record")
        st = json.loads(a.structured_json or "{}")
        eng = MS.engagement_full(s, a.engagement_id) if a.engagement_id else {}
        base = eng.get("base", {})
        # documents/artefacts tagged to this vendor/engagement
        arts = []
        for art in s.scalars(select(ArtefactRecord).where(
                ArtefactRecord.vendor_id == a.vendor_id)).all():
            arts.append({"artefact_id": art.artefact_id, "kind": art.artefact_type,
                         "title": art.name, "status": art.status,
                         "expiry_date": art.expiry_date,
                         "doc_link": getattr(art, "doc_link", None)})
        # controls assessed: DDQ / per-stage controls captured in structured snapshot
        controls = st.get("per_stage", [])
        can_approve = (getattr(u.role, "key", None) in ("admin", "vrm")) and not a.locked
        # IRQ (inherent-risk questionnaire) and Due Diligence question review
        irq = st.get("extracted_irq") or st.get("irq") or {}
        irq_rows = [{"question": k.replace("_", " ").title(), "answer": v}
                    for k, v in irq.items()] if isinstance(irq, dict) else []
        dd_rows = []
        for c in controls:
            for t in (c.get("turns", []) if isinstance(c, dict) else []):
                dd_rows.append({"area": (c.get("name") or "Control"),
                                "by": t.get("agent") or t.get("role"),
                                "detail": (t.get("body") or t.get("excerpt") or "")})
        for g in st.get("gaps", []):
            dd_rows.append({"area": g.get("domain") or "Gap", "by": "Assessment",
                            "detail": g.get("issue") or "", "resolution": g.get("resolution")})
        # all documents available / referred during the assessment (artefacts + stored docs)
        all_docs = list(arts)
        try:
            from app.features.lifecycle.documents import StoredDocument as _SD
            for d in s.scalars(select(_SD).where(
                    _SD.vendor_id == a.vendor_id)).all():
                all_docs.append({"artefact_id": d.doc_id, "title": d.filename,
                                 "kind": d.purpose or d.content_type, "status": "on file",
                                 "doc_link": f"/api/v2/documents/{d.doc_id}"})
        except Exception as _e:
            _obs_swallow('bro_app.py', _e)
        # findings linked to this assessment
        link_findings = [{"finding_id": f.finding_id, "title": f.title, "severity": f.severity,
                          "status": f.status, "source": f.source}
                         for f in s.scalars(select(FindingRecord).where(
                             FindingRecord.assessment_id == a.assessment_id)).all()]
        return {
            "assessment_id": a.assessment_id, "engagement_id": a.engagement_id,
            "vendor_id": a.vendor_id, "status": a.status, "locked": a.locked,
            "assigned_assessor": a.assessor_user,
            "irq": irq_rows, "due_diligence": dd_rows,
            "all_documents": all_docs, "linked_findings": link_findings,
            "can_assign": getattr(u.role, "key", None) in ("admin", "controller"),
            "outcome": a.outcome, "assessor_signed_off": a.assessor_signed_off,
            "scope": {"title": base.get("title"), "service_description": base.get("service_description"),
                      "data_classification": (eng.get("ext", {}) or {}).get("data_classification"),
                      "is_critical": base.get("is_critical")},
            "inherent": {"band": a.inherent_band, "detail": st.get("inherent_detail"),
                         "risks": st.get("risks", [])},
            "controls_assessed": controls,
            "documents": arts,
            "residual": {"band": a.residual_band, "verdict": st.get("verdict"),
                         "recommendation": st.get("recommendation")},
            "gaps": st.get("gaps", []),
            "transcript_available": bool(st.get("transcript")),
            "can_approve": can_approve,
        }

    @app.get("/api/v2/sectors")
    def v2_sectors(u: User = Depends(require("intel.financial"))):
        return FIN.SECTORS

    @app.post("/api/v2/research/web")
    def v2_web_research(b: FinResearchIn, s: Session = Depends(db),
                        u: User = Depends(require("intel.financial"))):
        """Live internet-grounded combined FDD + reputation research."""
        from app.features.intelligence import entity_resolve as ER
        res = ER.web_research_fdd_reputation(b.company, b.jurisdiction or "UK",
                                             b.identifier or "")
        audit(s, "v2.web_research", u.username,
              {"company": b.company, "matched": res.get("matched"),
               "available": res.get("available")})
        s.commit()
        return res

    # ===== Methodology library (admin-only) =====
    @app.post("/api/v2/research/fdd")
    def v2_research_fdd(b: AIResearchIn, s: Session = Depends(db),
                        u: User = Depends(require("intel.financial"))):
        if not _ai_live():
            return {"available": False, "holding": True, "message": AI_HOLDING}
        job_id = _start_research_job(vendor_id=b.vendor_id, company=b.company,
                                     jurisdiction=b.jurisdiction, identifier=b.identifier,
                                     mode="fdd", deep=bool(getattr(b, "deep", False)),
                                     actor=u.username)
        audit(s, "v2.research_fdd.start", u.username,
              {"vendor_id": b.vendor_id, "company": b.company, "job_id": job_id}); s.commit()
        return {"pending": True, "job_id": job_id,
                "message": ("Research started — running on the server. It will appear here when "
                            "complete and is filed in AI Reports even if you navigate away.")}

    @app.get("/api/v2/research/status/{job_id}")
    def v2_research_status(job_id: str, u: User = Depends(require("engagement.view"))):
        """Poll a background FDD/reputation research job. Returns the full result on
        completion (same shape as the old synchronous response)."""
        j = _research_job_status(job_id)
        if not j or (j.get("actor") and j["actor"] != u.username):
            return {"status": "unknown"}
        if j["status"] == "done":
            r = dict(j.get("result") or {}); r["status"] = "done"; return r
        if j["status"] == "error":
            return {"status": "error", "available": True, "holding": True,
                    "message": ("AI research did not complete: " + (j.get("error") or "unknown error")
                                + ". Check the AI provider/key/model and web-search entitlement in "
                                  "Settings → AI, then retry.")}
        import time as _t
        return {"status": "running", "elapsed_s": int(_t.time() - j.get("started", 0))}

    @app.post("/api/v2/procurement/pr-pull")
    def v2_pr_pull(b: PRPullIn, s: Session = Depends(db),
                   u: User = Depends(require("engagement.view"))):
        """MOCK purchasing-system pull. In production this calls the procurement
        connector; here it returns a plausible pulled package for the demo."""
        from app.features.domain.registry_models import VendorRecord
        pr = (b.pr_number or "").strip() or "PR-DEMO"
        vs = list(s.scalars(select(VendorRecord)).all())
        v = vs[abs(hash(pr)) % len(vs)] if vs else None
        return {"pr_number": pr, "matched": bool(v),
                "vendor_id": v.vendor_id if v else None,
                "vendor_name": v.legal_name if v else "Unmatched supplier",
                "tier": v.tier if v else None,
                "scope": "Provision of cloud hosting and data-processing services; ~50,000 "
                         "records; multi-region; API integration; offshore support.",
                "annual_value": 250000,
                "documents": ["Proposal.pdf", "Quote.pdf", "System_Design.pdf",
                              "Vendor_email_thread.eml", "SOC2_Type_II.pdf"]}

    @app.post("/api/v2/engagements/similar")
    def v2_similar_engagements(b: SimilarIn, s: Session = Depends(db),
                               u: User = Depends(require("engagement.view"))):
        """Check the Vendor & Entity database for similar engagements (mock heuristic:
        vendor-name + scope-keyword overlap)."""
        from app.features.domain.registry_models import VendorRecord, EngagementRecord
        ent = (b.entity or "").lower()
        scope_toks = {t for t in (b.scope or "").lower().replace(",", " ").split() if len(t) > 3}
        vmap = {v.vendor_id: v for v in s.scalars(select(VendorRecord)).all()}
        matches = []
        for e in s.scalars(select(EngagementRecord)).all():
            v = vmap.get(e.vendor_id); vname = (v.legal_name.lower() if v else "")
            if ent and (ent in vname or vname in ent):
                name_match = 1.0
            elif ent and any(tok in vname for tok in ent.split() if len(tok) > 3):
                name_match = 0.6
            else:
                name_match = 0.0
            title_toks = {t for t in (e.title or "").lower().replace(",", " ").split() if len(t) > 3}
            overlap = (len(scope_toks & title_toks) / max(1, len(scope_toks))) if scope_toks else 0.0
            score = round(0.6 * name_match + 0.4 * overlap, 2)
            if score >= 0.4:
                matches.append({"engagement_id": e.engagement_id, "vendor_id": e.vendor_id,
                                "vendor_name": v.legal_name if v else e.vendor_id,
                                "title": e.title, "stage": e.stage, "status": e.status,
                                "inherent_band": e.inherent_band, "residual_band": e.residual_band,
                                "score": score})
        matches.sort(key=lambda m: m["score"], reverse=True)
        top = matches[:5]
        return {"matches": top, "very_similar": bool(top and top[0]["score"] >= 0.7),
                "best": top[0] if top else None}

    @app.get("/api/v2/fin-monitor")
    def v2_finmon_list(s: Session = Depends(db), u: User = Depends(require("intel.financial"))):
        from app.features.domain.registry_models import FinMonitorRecord
        rows = s.scalars(select(FinMonitorRecord).order_by(FinMonitorRecord.id)).all()
        return [{"id": r.id, "vendor_id": r.vendor_id, "entity_name": r.entity_name,
                 "last_signal": r.last_signal, "last_swept": r.last_swept,
                 "last_result": r.last_result} for r in rows]

    @app.post("/api/v2/fin-monitor")
    def v2_finmon_add(b: FinMonitorAddIn, s: Session = Depends(db),
                      u: User = Depends(require("intel.financial"))):
        from app.features.intelligence import entity_resolve as ER
        from app.features.domain.registry_models import FinMonitorRecord
        ent = ER.resolve_entity(s, vendor_id=b.vendor_id, other_name=b.other_name)
        if ent["vendor_name"] == "(unspecified)":
            raise HTTPException(400, "provide a vendor_id or other_name")
        row = FinMonitorRecord(vendor_id=ent["vendor_id"], entity_name=ent["vendor_name"])
        s.add(row); s.flush()
        audit(s, "v2.finmon_empanel", u.username, {"entity": ent["vendor_name"], "id": row.id})
        s.commit()
        return {"id": row.id, "vendor_id": row.vendor_id, "entity_name": row.entity_name}

    @app.delete("/api/v2/fin-monitor/{mid}")
    def v2_finmon_remove(mid: int, s: Session = Depends(db),
                         u: User = Depends(require("intel.financial"))):
        from app.features.domain.registry_models import FinMonitorRecord
        row = s.get(FinMonitorRecord, mid)
        if row:
            s.delete(row); s.commit()
        return {"deleted": True}

    @app.post("/api/v2/fin-monitor/sweep")
    def v2_finmon_sweep(b: FinMonitorSweepIn, s: Session = Depends(db),
                        u: User = Depends(require("intel.financial"))):
        from app.features.domain.registry_models import FinMonitorRecord
        from app.agents import llm_config
        import datetime as _dt
        targets = ([s.get(FinMonitorRecord, b.monitor_id)] if b.monitor_id
                   else s.scalars(select(FinMonitorRecord)).all())
        targets = [t for t in targets if t]
        ts = _dt.datetime.now(_dt.timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
        swept = 0
        for t in targets:
            if llm_config.is_enabled():
                text = llm_config.complete(
                    PROMPTS.resolve(s, "financial_monitor"),
                    f"Monitoring sweep for: {t.entity_name}", domain="finance")
                signal = "ok"
                if text:
                    low = text.lower()
                    signal = ("distress" if "signal=distress" in low else
                              "watch" if "signal=watch" in low else "ok")
                t.last_result = text or "No material findings."
                t.last_signal = signal
            else:
                t.last_result = ("Live monitoring needs an AI key. With a key set, Vera "
                                 "sweeps authoritative sources for financial-health signals, "
                                 "profit warnings, rating changes and distress indicators.")
                t.last_signal = "ok"
            t.last_swept = ts
            swept += 1
            # R1: reconcile the panel signal into the attribute time-series + risk profile
            if t.vendor_id:
                from app.features.domain import master_service as MS
                MS.persist_monitor_result(s, t.vendor_id, t.last_signal, t.last_result)
                MS.refresh_risk_profile(s, t.vendor_id)
        audit(s, "v2.finmon_sweep", u.username, {"swept": swept})
        s.commit()
        return {"swept": swept, "last_swept": ts,
                "ai_enabled": llm_config.is_enabled()}

    # ============================================================
    # REQ 1 — VENDOR MASTER
    # ============================================================
    @app.get("/api/v2/vendor-master/{vid}")
    def v2_vendor_master_get(vid: str, s: Session = Depends(db),
                             u: User = Depends(require("vendor.view"))):
        _RBAC.assert_object_visible(s, u, "vendor", vid)
        from app.features.domain import master_service as MS
        # banking visible only to admin or vendor.critical holders
        inc_bank = u.role.key == "admin" or "vendor.critical" in {p.key for p in u.role.permissions}
        data = MS.get_vendor_master(s, vid, include_bank=inc_bank)
        if not data:
            raise HTTPException(404, "vendor not found")
        return data

    @app.get("/api/v2/vendor/{vid}/linkage")
    def v2_vendor_linkage(vid: str, s: Session = Depends(db),
                          u: User = Depends(require("vendor.view"))):
        """Engagements (with Active/Inactive status), contracts and documents linked
        to a vendor — for the visible vendor↔engagement linkage in the master record."""
        _RBAC.assert_object_visible(s, u, "vendor", vid)
        from app.features.domain.registry_models import EngagementRecord
        from app.features.domain.master_ext import ContractRecord, EngagementExt
        from app.features.lifecycle.documents import StoredDocument
        import datetime as _dt
        now = _dt.datetime.now(_dt.timezone.utc)
        engs = s.scalars(select(EngagementRecord).where(EngagementRecord.vendor_id == vid)).all()
        ext = {e.engagement_id: e for e in s.scalars(select(EngagementExt).where(
            EngagementExt.engagement_id.in_([e.engagement_id for e in engs]))).all()} if engs else {}

        def _active(e):
            st = (e.status or "").lower(); stg = (e.stage or "").lower()
            if stg in ("terminate", "terminated", "exited") or st in ("terminated", "expired", "inactive", "closed"):
                return False
            x = ext.get(e.engagement_id)
            exp = getattr(x, "contract_end", None) or getattr(x, "expiry_date", None) if x else None
            if exp:
                try:
                    ed = exp if isinstance(exp, _dt.datetime) else _dt.datetime.fromisoformat(str(exp))
                    if ed.tzinfo is None: ed = ed.replace(tzinfo=_dt.timezone.utc)
                    if ed < now: return False
                except Exception as _e:
                    _obs_swallow('bro_app.py', _e)
            return True
        engagements = [{"engagement_id": e.engagement_id, "title": e.title, "stage": e.stage,
                        "status": e.status, "active": _active(e),
                        "inherent_band": e.inherent_band, "residual_band": e.residual_band}
                       for e in engs]
        contracts = [{"contract_id": c.contract_id, "type": c.contract_type, "status": c.status,
                      "doc_link": c.doc_link, "engagement_id": c.engagement_id}
                     for c in s.scalars(select(ContractRecord).where(ContractRecord.vendor_id == vid)).all()]
        docs = [{"doc_id": d.doc_id, "filename": d.filename, "purpose": d.purpose,
                 "url": f"/api/v2/documents/{d.doc_id}"}
                for d in s.scalars(select(StoredDocument).where(StoredDocument.vendor_id == vid)
                                   .order_by(StoredDocument.id.desc())).all()]
        return {"vendor_id": vid, "engagements": engagements, "contracts": contracts,
                "documents": docs,
                "active_count": sum(1 for e in engagements if e["active"]),
                "inactive_count": sum(1 for e in engagements if not e["active"])}

    @app.put("/api/v2/vendor-master/{vid}")
    def v2_vendor_master_put(vid: str, b: VendorMasterIn, s: Session = Depends(db),
                             u: User = Depends(require("vendor.edit"))):
        from app.features.domain import master_service as MS
        v = s.scalars(select(VendorRecord).where(VendorRecord.vendor_id == vid)).first()
        if not v:
            raise HTTPException(404, "vendor not found")
        err = _validate_typed_fields(b.data)
        if err:
            raise HTTPException(422, err)
        inc_bank = b.include_bank and (
            u.role.key == "admin" or "vendor.critical" in {p.key for p in u.role.permissions})
        if b.include_bank and not inc_bank:
            raise HTTPException(403, "banking fields require elevated permission")
        MS.update_vendor_master(s, vid, b.data, include_bank=inc_bank)
        from app.features.domain import registry_service as _RS
        n_re = _RS.schedule_reassessment(s, vendor_id=vid, reason="Vendor master data updated")
        audit(s, "v2.vendor_master_update", u.username,
              {"vendor_id": vid, "bank": inc_bank, "reassessment_scheduled": n_re})
        s.commit()
        return MS.get_vendor_master(s, vid, include_bank=inc_bank)

    # ============================================================
    # REQ 2 — VENDOR ATTRIBUTE DATABASE
    # ============================================================
    @app.get("/api/v2/vendor-attributes/{vid}")
    def v2_vendor_attributes(vid: str, s: Session = Depends(db),
                             u: User = Depends(require("vendor.view"))):
        _RBAC.assert_object_visible(s, u, "vendor", vid)
        from app.features.domain import master_service as MS
        v = s.scalars(select(VendorRecord).where(VendorRecord.vendor_id == vid)).first()
        if not v:
            raise HTTPException(404, "vendor not found")
        return MS.vendor_attributes(s, vid)

    @app.post("/api/v2/vendor-attributes/{vid}/domain/{domain}")
    def v2_attr_domain(vid: str, domain: str, b: AttrDomainIn, s: Session = Depends(db),
                       u: User = Depends(require("vendor.edit"))):
        from app.features.domain import master_service as MS
        fn = {"privacy": MS.update_privacy, "cyber": MS.update_cyber,
              "resilience": MS.update_resilience, "esg": MS.update_esg,
              "governance": MS.update_governance}.get(domain)
        if not fn:
            raise HTTPException(404, "unknown attribute domain")
        fn(s, vid, b.data)
        audit(s, "v2.attr_update", u.username, {"vendor_id": vid, "domain": domain})
        s.commit()
        return MS.vendor_attributes(s, vid)

    @app.post("/api/v2/vendor-attributes/{vid}/insurance")
    def v2_attr_insurance(vid: str, b: InsuranceIn, s: Session = Depends(db),
                          u: User = Depends(require("vendor.edit"))):
        from app.features.domain import master_service as MS
        MS.add_insurance(s, vid, b.model_dump())
        audit(s, "v2.insurance_add", u.username, {"vendor_id": vid, "type": b.policy_type})
        s.commit()
        return MS.vendor_attributes(s, vid)

    @app.post("/api/v2/vendor-attributes/{vid}/monitor-signal")
    def v2_attr_signal(vid: str, b: MonitorSignalIn, s: Session = Depends(db),
                       u: User = Depends(require("vendor.edit"))):
        from app.features.domain import master_service as MS
        MS.add_monitor_signal(s, vid, b.signal_type, b.value, b.source)
        s.commit()
        return {"ok": True}

    @app.post("/api/v2/vendor-attributes/{vid}/refresh-rollups")
    def v2_attr_refresh(vid: str, s: Session = Depends(db),
                        u: User = Depends(require("vendor.view"))):
        _RBAC.assert_object_visible(s, u, "vendor", vid)
        from app.features.domain import master_service as MS
        MS.refresh_cyber_certs(s, vid)
        rp = MS.refresh_risk_profile(s, vid)
        s.commit()
        return {"refreshed": True, "inherent_band": rp.inherent_band,
                "residual_band": rp.residual_band, "open_findings": rp.open_findings}

    # ============================================================
    # REQ 3 — ENGAGEMENT REGISTER
    # ============================================================
    @app.get("/api/v2/engagement-register/{eid}")
    def v2_eng_full(eid: str, s: Session = Depends(db),
                    u: User = Depends(require("engagement.view"))):
        from app.features.domain import master_service as MS
        data = MS.engagement_full(s, eid)
        if not data:
            raise HTTPException(404, "engagement not found")
        return data

    @app.get("/api/v2/engagement/{eid}/irq-dd")
    def v2_eng_irqdd(eid: str, s: Session = Depends(db),
                     u: User = Depends(require("engagement.view"))):
        """Full IRQ + Due-Diligence detail for an engagement, joined to question
        text, with scores, bands and linked documents — reviewable to answer level."""
        import json as _json
        from app.features.domain.registry_models import AssessmentRecord, EngagementRecord, VendorRecord
        from app.features.lifecycle.documents import StoredDocument
        from app.features.domain.master_ext import ContractRecord
        from app.features.assessment.bro_engine import IRQ_QUESTIONS, DDQ_DOMAINS
        eng = s.query(EngagementRecord).filter(EngagementRecord.engagement_id == eid).first()
        if not eng:
            raise HTTPException(404, "engagement not found")
        a = (s.query(AssessmentRecord)
             .filter(AssessmentRecord.engagement_id == eid)
             .order_by(AssessmentRecord.id.desc()).first())
        ven = s.query(VendorRecord).filter(VendorRecord.vendor_id == eng.vendor_id).first()
        d = {}
        if a and a.structured_json:
            try: d = _json.loads(a.structured_json)
            except Exception: d = {}
        irq_ans = d.get("irq", {}) or {}
        ddq_ans = d.get("ddq", {}) or {}

        def fmt(v):
            return ", ".join(v) if isinstance(v, list) else ("" if v is None else str(v))
        irq_rows = [{"id": q["id"], "question": q["text"], "type": q["type"],
                     "answer": fmt(irq_ans.get(q["id"])),
                     "answered": irq_ans.get(q["id"]) not in (None, "", [])}
                    for q in IRQ_QUESTIONS]
        ddq_domains = []
        for dom in DDQ_DOMAINS:
            qs = [{"id": q["id"], "question": q["text"], "critical": q.get("critical", False),
                   "response": ddq_ans.get(q["id"], "—")} for q in dom["questions"]]
            ddq_domains.append({"id": dom["id"], "name": dom["name"], "questions": qs})

        docs = (s.query(StoredDocument)
                .filter((StoredDocument.engagement_id == eid) | (StoredDocument.vendor_id == eng.vendor_id))
                .order_by(StoredDocument.id.desc()).all())
        documents = [{"doc_id": x.doc_id, "filename": x.filename, "purpose": x.purpose,
                      "scope": "engagement" if x.engagement_id == eid else "vendor",
                      "url": f"/api/v2/documents/{x.doc_id}"} for x in docs]
        contracts = [{"contract_id": c.contract_id, "type": c.contract_type, "status": c.status,
                      "doc_link": c.doc_link}
                     for c in s.query(ContractRecord)
                     .filter((ContractRecord.engagement_id == eid) | (ContractRecord.vendor_id == eng.vendor_id)).all()]
        return {
            "engagement_id": eid, "title": eng.title, "vendor_id": eng.vendor_id,
            "vendor_name": ven.legal_name if ven else eng.vendor_id,
            "assessment_id": a.assessment_id if a else None,
            "status": a.status if a else None, "outcome": a.outcome if a else None,
            "assessor": a.assessor_user if a else None,
            "tier": d.get("tier"), "schema": d.get("schema"),
            "inherent_band": d.get("inherent_band", eng.inherent_band),
            "residual": d.get("residual", {"band": eng.residual_band}),
            "weighted_pct": d.get("weighted_pct"), "completeness_cls": d.get("completeness_cls"),
            "domain_scores": d.get("domain_scores", {}),
            "recommendation": d.get("recommendation"), "route": d.get("route"),
            "scope_summary": d.get("scope_summary"),
            "irq": irq_rows, "ddq": ddq_domains,
            "documents": documents, "contracts": contracts,
            "has_assessment": bool(a),
        }

    # ===== Exit Strategy & Exit Planning (vendor level, CMORG-aligned) =====
    @app.put("/api/v2/engagement-register/{eid}")
    def v2_eng_ext_put(eid: str, b: EngExtIn, s: Session = Depends(db),
                       u: User = Depends(require("engagement.edit"))):
        from app.features.domain import master_service as MS
        eng = s.scalars(select(EngagementRecord).where(
            EngagementRecord.engagement_id == eid)).first()
        if not eng:
            raise HTTPException(404, "engagement not found")
        # base-record fields (inherent/residual band) persist to the engagement itself
        data = dict(b.data or {})
        for base_field in ("inherent_band", "residual_band"):
            if base_field in data:
                setattr(eng, base_field, data.pop(base_field))
        MS.update_eng_ext(s, eid, data)
        from app.features.domain import registry_service as _RS
        _RS.schedule_reassessment(s, engagement_id=eid, reason="Engagement data updated")
        audit(s, "v2.engagement_update", u.username, {"engagement_id": eid})
        s.commit()
        return MS.engagement_full(s, eid)

    @app.get("/api/v2/engagement/{eid}/assessments")
    def v2_engagement_assessments(eid: str, s: Session = Depends(db),
                                  u: User = Depends(require("engagement.view"))):
        from datetime import datetime, date
        from app.features.domain.registry_models import EngagementRecord, AssessmentRecord
        e = s.scalars(select(EngagementRecord).where(
            EngagementRecord.engagement_id == eid)).first()
        if not e:
            raise HTTPException(404, "engagement not found")
        asms = list(s.scalars(select(AssessmentRecord).where(
            AssessmentRecord.engagement_id == eid)).all())
        asms.sort(key=lambda a: (a.created_at or datetime.min), reverse=True)  # latest first
        rows = [{"assessment_id": a.assessment_id, "status": a.status,
                 "inherent_band": a.inherent_band, "residual_band": a.residual_band,
                 "outcome": a.outcome, "assessor": a.assessor_user,
                 "signed_off": a.assessor_signed_off, "locked": a.locked,
                 "date": a.created_at.isoformat() if a.created_at else None} for a in asms]
        due = e.next_assessment_due
        overdue = bool(due and due <= date.today().isoformat())
        return {"engagement_id": eid, "last_assessment_date": e.last_assessment_date,
                "next_assessment_due": due, "reassessment_due": overdue,
                "reassessment_reason": e.reassessment_reason, "count": len(rows),
                "assessments": rows}

    @app.get("/api/v2/reports/expired-assessments")
    def v2_expired_assessments(s: Session = Depends(db),
                               u: User = Depends(require("engagement.view"))):
        """Engagements whose reassessment is overdue (next_assessment_due < today),
        most-overdue first. Powers the Management → Expired Assessments report."""
        from datetime import date
        from app.features.domain.registry_models import EngagementRecord, VendorRecord
        today = date.today()
        today_iso = today.isoformat()
        vmap = {v.vendor_id: v.legal_name for v in s.scalars(select(VendorRecord)).all()}
        rows = []
        total = 0
        for e in s.scalars(select(EngagementRecord)).all():
            if not e.next_assessment_due:
                continue
            total += 1
            if e.next_assessment_due >= today_iso:
                continue
            try:
                y, m, d = (int(x) for x in e.next_assessment_due.split("-"))
                days_overdue = (today - date(y, m, d)).days
            except Exception:
                days_overdue = None
            rows.append({"engagement_id": e.engagement_id, "vendor_id": e.vendor_id,
                         "vendor": vmap.get(e.vendor_id, e.vendor_id),
                         "title": e.title, "inherent_band": e.inherent_band,
                         "residual_band": e.residual_band,
                         "is_critical": getattr(e, "is_critical", False),
                         "last_assessment_date": e.last_assessment_date,
                         "next_assessment_due": e.next_assessment_due,
                         "days_overdue": days_overdue,
                         "owner": e.owner_user, "assessor": e.assigned_assessor,
                         "reassessment_reason": e.reassessment_reason})
        rows.sort(key=lambda r: (r["days_overdue"] is None, -(r["days_overdue"] or 0)))
        by_band = {}
        for r in rows:
            by_band[r["inherent_band"] or "—"] = by_band.get(r["inherent_band"] or "—", 0) + 1
        return {"as_of": today_iso, "expired": len(rows),
                "total_with_due": total, "by_band": by_band, "rows": rows}

    # ---------------- PESTLE threat intelligence ----------------
    @app.get("/api/v2/format-settings")
    def v2_format_settings(s: Session = Depends(db), u: User = Depends(actor)):
        from app.features.domain import config_store as CFG
        return {"date": CFG.get_config(s, "format.date", "MM-DD-YYYY"),
                "currency": CFG.get_config(s, "format.currency", "USD")}

    @app.post("/api/v2/engagement-register/{eid}/child")
    def v2_eng_child_add(eid: str, b: EngChildIn, s: Session = Depends(db),
                         u: User = Depends(require("engagement.edit"))):
        from app.features.domain import master_service as MS
        if b.kind not in ("deliverable", "milestone", "sla", "obligation", "personnel"):
            raise HTTPException(400, "invalid child kind")
        row = MS.add_eng_child(s, eid, b.kind, b.data)
        audit(s, "v2.engagement_child_add", u.username, {"engagement_id": eid, "kind": b.kind})
        s.commit()
        return {"id": row.id, "kind": b.kind}

    @app.delete("/api/v2/engagement-register/{eid}/child/{kind}/{cid}")
    def v2_eng_child_del(eid: str, kind: str, cid: int, s: Session = Depends(db),
                         u: User = Depends(require("engagement.edit"))):
        from app.features.domain import master_ext as MX
        model = {"deliverable": MX.EngagementDeliverable, "milestone": MX.EngagementMilestone,
                 "sla": MX.EngagementSLA, "obligation": MX.EngagementObligation,
                 "personnel": MX.EngagementPersonnel}.get(kind)
        if not model:
            raise HTTPException(400, "invalid child kind")
        row = s.get(model, cid)
        if row and row.engagement_id == eid:
            s.delete(row); s.commit()
        return {"deleted": True}

    @app.get("/api/v2/obligations/overdue")
    def v2_obligations_overdue(s: Session = Depends(db),
                               u: User = Depends(require("engagement.view"))):
        from app.features.domain import master_service as MS
        return MS.overdue_obligations(s)

    # ============================================================
    # REQ 2 — CONTRACT ENTITY
    # ============================================================
    @app.post("/api/v2/engagement-register/{eid}/sync-contract")
    def v2_eng_sync_contract(eid: str, s: Session = Depends(db),
                             u: User = Depends(require("engagement.edit"))):
        from app.features.domain import master_service as MS
        row = MS.sync_engagement_contract(s, eid)
        s.commit()
        if not row:
            return {"synced": False, "reason": "no contract_reference on engagement"}
        return {"synced": True, "contract_id": row.contract_id}

    # ============================================================
    # REQ 3 — CRITICAL VENDORS MODULE
    # ============================================================
    @app.post("/api/v2/critical-vendors/analyse")
    def v2_crit_analyse(b: CriticalAnalysisIn, s: Session = Depends(db),
                        u: User = Depends(require("vendor.critical"))):
        from app.features.domain import master_service as MS
        res = MS.run_critical_analysis(s, b.vendor_id)
        audit(s, "v2.critical_analysis", u.username,
              {"analysed": res["analysed"], "critical": len(res["critical_vendors"])})
        s.commit()
        return res

    @app.get("/api/v2/critical-vendors")
    def v2_crit_list(s: Session = Depends(db), u: User = Depends(require("vendor.view"))):
        from app.features.domain import master_service as MS
        return MS.list_critical(s)

    @app.post("/api/v2/critical-vendors/{vid}/override")
    def v2_crit_override(vid: str, b: CriticalOverrideIn, s: Session = Depends(db),
                         u: User = Depends(require("vendor.critical"))):
        from app.features.domain import master_service as MS
        res = MS.override_vendor_criticality(s, vid, b.is_critical, b.reason, u.username)
        audit(s, "v2.criticality_override", u.username,
              {"vendor_id": vid, "is_critical": b.is_critical})
        s.commit()
        return res

    @app.get("/api/v2/slas")
    def v2_list_slas(engagement_id: Optional[str] = None, vendor_id: Optional[str] = None,
                     s: Session = Depends(db), u: User = Depends(require("engagement.view"))):
        return [PERF.sla_row(s, x) for x in PERF.list_slas(s, engagement_id, vendor_id)]

    @app.post("/api/v2/slas")
    def v2_create_sla(b: SLAIn, s: Session = Depends(db),
                      u: User = Depends(require("engagement.edit"))):
        sla = PERF.create_sla(s, engagement_id=b.engagement_id, vendor_id=b.vendor_id,
                              description=b.description, threshold_type=b.threshold_type or "min",
                              threshold=b.threshold, unit=b.unit or "", baseline=b.baseline,
                              window=b.window or "monthly", source=b.source or "manual",
                              contract_id=b.contract_id, created_by=u.username)
        audit(s, "v2.sla_created", u.username,
              {"sla_id": sla.sla_id, "engagement_id": b.engagement_id})
        s.commit()
        return PERF.sla_row(s, sla)

    @app.put("/api/v2/slas/{sla_id}")
    def v2_update_sla(sla_id: str, b: SLAEditIn, s: Session = Depends(db),
                      u: User = Depends(require("engagement.edit"))):
        sla = s.scalar(select(PERF.SLARecord).where(PERF.SLARecord.sla_id == sla_id))
        if not sla:
            raise HTTPException(404, "SLA not found")
        PERF.update_sla(s, sla, **b.model_dump(exclude_none=True))
        audit(s, "v2.sla_updated", u.username, {"sla_id": sla_id})
        s.commit()
        return PERF.sla_row(s, sla)

    @app.delete("/api/v2/slas/{sla_id}")
    def v2_delete_sla(sla_id: str, s: Session = Depends(db),
                      u: User = Depends(require("engagement.edit"))):
        sla = s.scalar(select(PERF.SLARecord).where(PERF.SLARecord.sla_id == sla_id))
        if not sla:
            raise HTTPException(404, "SLA not found")
        sla.active = False
        audit(s, "v2.sla_deleted", u.username, {"sla_id": sla_id})
        s.commit()
        return {"sla_id": sla_id, "deleted": True}

    @app.post("/api/v2/slas/{sla_id}/measurements")
    def v2_sla_measure(sla_id: str, b: MeasurementIn, s: Session = Depends(db),
                       u: User = Depends(require("engagement.edit"))):
        sla = s.scalar(select(PERF.SLARecord).where(PERF.SLARecord.sla_id == sla_id))
        if not sla:
            raise HTTPException(404, "SLA not found")
        PERF.record_measurement(s, sla_id, b.period, b.value, u.username)
        audit(s, "v2.sla_measurement", u.username,
              {"sla_id": sla_id, "period": b.period, "value": b.value})
        s.commit()
        return PERF.sla_row(s, sla)

    @app.post("/api/v2/slas/extract")
    def v2_sla_extract(b: ExtractIn, s: Session = Depends(db),
                       u: User = Depends(require("engagement.edit"))):
        created = PERF.extract_slas(s, engagement_id=b.engagement_id, vendor_id=b.vendor_id,
                                    mode=b.mode or "contract", contract_id=b.contract_id,
                                    created_by=u.username)
        audit(s, "v2.sla_extract", u.username,
              {"engagement_id": b.engagement_id, "mode": b.mode, "count": len(created)})
        s.commit()
        return {"extracted": len(created), "slas": [PERF.sla_row(s, x) for x in created]}

    @app.get("/api/v2/slas/summary")
    def v2_sla_summary(engagement_id: str, s: Session = Depends(db),
                       u: User = Depends(require("engagement.view"))):
        return PERF.analyse(s, engagement_id)

    @app.post("/api/v2/slas/enquiry")
    def v2_sla_enquiry(b: EnquiryIn, s: Session = Depends(db),
                       u: User = Depends(require("engagement.view"))):
        return PERF.enquire(s, b.engagement_id, b.question)

    # ================= PERFORMANCE ISSUES (mirrors risk register) =================
    @app.get("/api/v2/engagements/{eid}/assessment-report")
    def v2_engagement_assessment_report(eid: str, s: Session = Depends(db),
                                        u: User = Depends(require("engagement.view"))):
        _RBAC.assert_object_visible(s, u, 'engagement', eid)
        """A detailed, structured assessment report for an entire engagement:
        the vendor, the engagement, every assessment, findings, documents and
        a decision summary — the basis for the Vendor 360 report view."""
        from app.features.domain import registry_models as RM
        from app.features.lifecycle import documents as DOCS
        eng = s.scalar(select(RM.EngagementRecord).where(RM.EngagementRecord.engagement_id == eid))
        if not eng:
            raise HTTPException(404, "engagement not found")
        ven = None
        if eng.vendor_id:
            ven = s.scalar(select(RM.VendorRecord).where(RM.VendorRecord.vendor_id == eng.vendor_id))
        assessments = s.scalars(select(RM.AssessmentRecord)
                                .where(RM.AssessmentRecord.engagement_id == eid)
                                .order_by(RM.AssessmentRecord.id.desc())).all()
        findings = s.scalars(select(RM.FindingRecord)
                             .where(RM.FindingRecord.engagement_id == eid)).all()
        docs = s.scalars(select(DOCS.StoredDocument)
                         .where(DOCS.StoredDocument.engagement_id == eid)).all()
        latest = assessments[0] if assessments else None
        sev_order = {"Critical": 0, "High": 1, "Medium": 2, "Low": 3}
        fnd = sorted(findings, key=lambda f: sev_order.get(f.severity, 9))
        return {
            "engagement": {"engagement_id": eng.engagement_id, "title": eng.title,
                           "status": getattr(eng, "status", None),
                           "owner": getattr(eng, "engagement_owner", None),
                           "service": getattr(eng, "service_description", None)},
            "vendor": ({"vendor_id": ven.vendor_id,
                        "name": ven.trading_name or ven.legal_name,
                        "criticality": getattr(ven, "criticality", None),
                        "country": getattr(ven, "incorporation_country", None)} if ven else None),
            "latest_assessment": ({
                "assessment_id": latest.assessment_id, "status": latest.status,
                "inherent_band": latest.inherent_band, "residual_band": latest.residual_band,
                "outcome": latest.outcome, "assessor": latest.assessor_user,
                "signed_off": latest.assessor_signed_off,
                "created_at": latest.created_at.isoformat() if latest.created_at else None,
            } if latest else None),
            "assessment_history": [{
                "assessment_id": a.assessment_id, "status": a.status,
                "inherent_band": a.inherent_band, "residual_band": a.residual_band,
                "outcome": a.outcome,
                "created_at": a.created_at.isoformat() if a.created_at else None,
            } for a in assessments],
            "findings": [{
                "finding_id": f.finding_id, "title": f.title, "severity": f.severity,
                "domain": getattr(f, "domain", None), "status": f.status,
                "owner": getattr(f, "owner", None),
            } for f in fnd],
            "findings_summary": {sev: sum(1 for f in findings if f.severity == sev)
                                 for sev in ["Critical", "High", "Medium", "Low"]},
            "documents": [{
                "doc_id": d.doc_id, "filename": d.filename, "purpose": d.purpose,
                "size_bytes": d.size_bytes, "uploaded_by": d.uploaded_by,
                "created_at": d.created_at.isoformat() if d.created_at else None,
            } for d in docs],
            "counts": {"assessments": len(assessments), "findings": len(findings),
                       "documents": len(docs)},
        }

    # ================= BRO CHAT — INTERIM AI REPORT (PDF) =================
    @app.get("/api/v2/vendor360/portfolio")
    def v2_vendor360_portfolio(s: Session = Depends(db),
                               u: User = Depends(require("vendor.view"))):
        from app.features.domain import master_service as MS
        rows = MS.vendor360_portfolio(s)
        # Row-level isolation: supplier sees only itself, buyer only its BU estate.
        _allowed = _RBAC.scoped_vendor_ids(s, u)
        if _allowed is not None:
            rows = [r for r in rows
                    if isinstance(r, dict) and r.get("vendor_id") in _allowed]
        return rows

    @app.get("/api/v2/vendor360/{vid}")
    def v2_vendor360(vid: str, s: Session = Depends(db),
                     u: User = Depends(require("vendor.view"))):
        _RBAC.assert_object_visible(s, u, "vendor", vid)
        from app.features.domain import master_service as MS
        data = MS.vendor360(s, vid)
        if not data:
            raise HTTPException(404, "vendor not found")
        s.commit()  # vendor360 refreshes the risk-profile snapshot
        return data


    return r
